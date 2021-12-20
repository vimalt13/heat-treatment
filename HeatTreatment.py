import dash
import dash_core_components as dcc
import dash_html_components as html
import dash_bootstrap_components as dbc
from dash.dependencies import Input, Output, State
from dash.dash import no_update
from dash_extensions import Download
from dash_extensions.snippets import send_data_frame
from dash_extensions.snippets import send_file
from flask import Flask
from waitress import serve
import dash_table
import numpy as np
import pandas as pd
from plotly.subplots import make_subplots
import plotly.graph_objects as go
from sklearn.ensemble import RandomForestRegressor
import joblib
import plotly.express as px
import base64
#from datetime import date
import openpyxl as opxl
from openpyxl.utils.cell import get_column_letter
import xlrd 
import xlwt
from xlwt import Workbook 
import xlsxwriter
import os
from datetime import datetime
import re
import psycopg2
import warnings
from flask_sqlalchemy import SQLAlchemy
warnings.filterwarnings('ignore')
from datetime import datetime
#import timedelta
from datetime import timedelta 

#import local functions and models
import configparser
config = configparser.ConfigParser()
config.read('config/settings.ini')
import functions
import functions2
import model_vtz

path = 'doc/VTZ_pipes.xlsx'
df=pd.read_excel(path, sheet_name='model')
list_marki=pd.read_excel(path, sheet_name='marki')
list_kalibr=pd.read_excel(path, sheet_name='kalibr')
data_prop=pd.read_excel('doc/data_properties_VTZ.xlsx',sheet_name='скорости_все')
df_for_model=pd.read_excel(path, sheet_name='match_with_model')
except_model=pd.read_excel('doc/except.xlsx')

def ff(x):
    return list(x)[0]

list_marki_1 = list_marki['marki'].unique()
list_kalibr_1 = list_kalibr['kalibr'].unique()

#TABLE FOR SELECT ID
def choose_table_create(df_sql):
    
    if len(df_sql.columns) >1:
        tb = df_sql[['scenarioid', 'datetime', 'vtz_report1_1', 'vtz_report1_12', 
                             'vtz_report1_3', 'vtz_report2_1', 'vtz_report2_3']]
        try:
            tb['datetime'] = tb['datetime'].astype('datetime64[ns]').dt.round('S').dt.strftime('%Y-%m-%d %H:%M:%S')
            tb['datetime'] = tb['datetime'].astype('string')
        except:
            tb['datetime'] = tb['datetime']
    else:
        tb = df_sql
        tb['Дата создания'] =  None
        tb['Калибр, мм'] =  None
        tb['Код отправки'] =  None
        tb['Марка стали'] =  None
        tb['Диаметр трубы, мм'] =  None
        tb['Толщина стенки, мм'] =  None
        
    return tb

head = functions2.sql_data_LIMIT500(order_id = 'datetime', limit = 5).loc[0].T #default table - 'data_report'
head = pd.DataFrame(head)
head = head.reset_index()
head.rename(columns = {'index':'name', 0:'vals'}, inplace = True)

###первая таблица
def display_tb(df, df2):
    df1 = df.copy()

    if df2.empty == True:
        tb = df[df['num'].isin(range(9))]
        tb = tb.sort_values(by=['num'])
        line = pd.DataFrame({"name_properties_2": 'Заготовка', "Значение": None}, index=[5])
        tb = pd.concat([tb.iloc[:5], line, tb.iloc[5:]]).reset_index(drop=True)
        tb = tb.drop(4)
        tb['Значение'] = None
        
        tb2 = df[df['num'].isin(range(9,19))]
        tb2 = tb2.sort_values(by=['num'])
        tb2['Значение'] = None
        
        tb3 = df[df['num'].isin(range(19,41))]
        tb3 = tb3.sort_values(by=['num'])
        tb3['Значение'] = None
        
        tb4 = df[df['num'].isin(range(41,68))]
        tb4 = tb4.sort_values(by=['num'])
        tb4['Значение'] = None      
        
    else:
        df2 = df2.T
        df2 = df2.reset_index()
        df2.rename(columns = {'index': 'Report', 0: 'Значение'}, inplace = True)
        df2['Report'] = df2.Report.replace({'vtz':'VTZ'}, regex=True)

        df3 = df1.merge(df2, left_on='Report', right_on='Report', how = 'left')
        val = df3['Значение'][4].split(sep = '/')
        df3['Значение'][0] = val[1]
        df3['Значение'][1] = val[0]

        tb = df3[df3['num'].isin(range(9))]
        tb = tb.sort_values(by=['num'])
        line = pd.DataFrame({"name_properties_2": 'Заготовка', "Значение": None}, index=[5])
        tb = pd.concat([tb.iloc[:5], line, tb.iloc[5:]]).reset_index(drop=True)
        
        tb2 = df3[df3['num'].isin(range(9,19))]
        tb2 = tb2.sort_values(by=['num'])
        
        tb3 = df3[df3['num'].isin(range(19,41))]
        tb3 = tb3.sort_values(by=['num'])
        
        tb4 = df3[df3['num'].isin(range(41,68))]
        tb4 = tb4.sort_values(by=['num'])

    tb = tb[['name_properties_2', 'Значение']]
    tb = tb.rename(columns = {'name_properties_2': 'Параметр', 'Значение': 'Значение'}, inplace = False)
    tb = tb.reset_index(drop=True)

    tb2 = tb2[['name_properties_2', 'Значение']]
    tb2 = tb2.rename(columns = {'name_properties_2': 'Параметр', 'Значение': 'Значение'}, inplace = False)
    tb2 = tb2.reset_index(drop=True)

    tb3 = tb3[['name_properties_2', 'Значение']]
    tb3 = tb3.rename(columns = {'name_properties_2': 'Параметр', 'Значение': 'Значение'}, inplace = False)
    tb3 = tb3.reset_index(drop=True)
    line = pd.DataFrame({"Параметр": 'Клеть #1', "Значение": None}, index=[0])
    tb3 = pd.concat([line, tb3]).reset_index(drop=True)
    tb3 = tb3.reset_index(drop=True)
    line = pd.DataFrame({"Параметр": 'Клеть #2', "Значение": None}, index=[4])
    tb3 = pd.concat([tb3.iloc[:4], line, tb3.iloc[4:]]).reset_index(drop=True)

    line = pd.DataFrame({"Параметр": 'Клеть #3', "Значение": None}, index=[8])
    tb3 = pd.concat([tb3.iloc[:8], line, tb3.iloc[8:]]).reset_index(drop=True)

    line = pd.DataFrame({"Параметр": 'Клеть #4', "Значение": None}, index=[12])
    tb3 = pd.concat([tb3.iloc[:12], line, tb3.iloc[12:]]).reset_index(drop=True)

    line = pd.DataFrame({"Параметр": 'Клеть #5', "Значение": None}, index=[16])
    tb3 = pd.concat([tb3.iloc[:16], line, tb3.iloc[16:]]).reset_index(drop=True)

    line = pd.DataFrame({"Параметр": 'Клеть #6', "Значение": None}, index=[20])
    tb3 = pd.concat([tb3.iloc[:20], line, tb3.iloc[20:]]).reset_index(drop=True)

    line = pd.DataFrame({"Параметр": 'Клеть #7', "Значение": None}, index=[24])
    tb3 = pd.concat([tb3.iloc[:24], line, tb3.iloc[24:]]).reset_index(drop=True)
    
    tb4 = tb4[['name_properties_2', 'Значение']]
    tb4 = tb4.rename(columns = {'name_properties_2': 'Параметр', 'Значение': 'Значение'}, inplace = False)
    tb4 = tb4.reset_index(drop=True)
    line = pd.DataFrame({"Параметр": 'Клеть #1', "Значение": None}, index=[0])
    tb4 = pd.concat([line, tb4]).reset_index(drop=True)
    tb4 = tb4.reset_index(drop=True)
    line = pd.DataFrame({"Параметр": 'Клеть #2', "Значение": None}, index=[4])
    tb4 = pd.concat([tb4.iloc[:4], line, tb4.iloc[4:]]).reset_index(drop=True)

    line = pd.DataFrame({"Параметр": 'Клеть #3', "Значение": None}, index=[8])
    tb4 = pd.concat([tb4.iloc[:8], line, tb4.iloc[8:]]).reset_index(drop=True)

    line = pd.DataFrame({"Параметр": 'Клеть #4', "Значение": None}, index=[12])
    tb4 = pd.concat([tb4.iloc[:12], line, tb4.iloc[12:]]).reset_index(drop=True)

    line = pd.DataFrame({"Параметр": 'Клеть #5', "Значение": None}, index=[16])
    tb4 = pd.concat([tb4.iloc[:16], line, tb4.iloc[16:]]).reset_index(drop=True)

    line = pd.DataFrame({"Параметр": 'Клеть #6', "Значение": None}, index=[20])
    tb4 = pd.concat([tb4.iloc[:20], line, tb4.iloc[20:]]).reset_index(drop=True)

    line = pd.DataFrame({"Параметр": 'Клеть #7', "Значение": None}, index=[24])
    tb4 = pd.concat([tb4.iloc[:24], line, tb4.iloc[24:]]).reset_index(drop=True)

    line = pd.DataFrame({"Параметр": 'Клеть #8', "Значение": None}, index=[28])
    tb4 = pd.concat([tb4.iloc[:28], line, tb4.iloc[28:]]).reset_index(drop=True)

    line = pd.DataFrame({"Параметр": 'Клеть #9', "Значение": None}, index=[32])
    tb4 = pd.concat([tb4.iloc[:32], line, tb4.iloc[32:]]).reset_index(drop=True)
    
    return tb, tb2, tb3, tb4

#функция для переноса данных в эксель
def template(replacementTextKeyPairs, card_name):

    book = opxl.load_workbook('./doc/' + card_name)
    ws = book.worksheets[0] 
    number_of_rows = ws.max_row
    number_of_cols = ws.max_column
    
    for i in range(number_of_cols):
        for k in range(number_of_rows):

            cellValue = str(ws[get_column_letter(i+1)+str(k+1)].value)

            for key in replacementTextKeyPairs.keys():

                if str(cellValue) == key:
                    newCellValue = replacementTextKeyPairs.get(key)
                    ws[get_column_letter(i+1)+str(k+1)] = str(newCellValue)

    book.save(card_name)
    return book.save(card_name)

###CREATE FIGURE _APP MODEL
def graph_1(x_data, y_data):
    #config={'responsive': True}
    fig=make_subplots(rows=1,cols=1)

    fig.update_layout(barmode='relative', 
                      #title_text='график трубы',
                      plot_bgcolor = '#ecf0f1',
                      #plot_bgcolor = '#B0C4DE',
                      paper_bgcolor = '#ecf0f1', 
                      yaxis_title ='Cкорость трубы в клети, мм/сек',
                      xaxis_title ='Номер клети',
                      #height='400',
                       autosize = True,
                           margin=dict(
                                l=20,
                                r=35,
                                b=5,
                                t=40,
                                pad=4),
                     )
    
    fig.add_scatter(x=x_data, y=y_data, name="исторические данные", mode='lines', 
                    line=dict(color='#191970', width=3),
                     row=1,col=1)

    
    return fig

def graph_2(x_data, y_data, z_data):
    fig=make_subplots(rows=1,cols=1)
    fig.update_layout(barmode='relative', 
                      #title_text='график трубы',
                      plot_bgcolor = '#ecf0f1',
                      #plot_bgcolor = '#B0C4DE',
                      paper_bgcolor = '#ecf0f1', 
                      yaxis_title ='Cкорость трубы в клети, мм/сек',
                      xaxis_title ='Номер клети',
                      #height='400',
                      showlegend = True,
                      legend=dict(
                            orientation="h",
                            yanchor="bottom",
                            y=1.1,
                            xanchor="left",
                            x=0.25
                        ),
                       autosize = True,
                           margin=dict(
                                l=20,
                                r=35,
                                b=5,
                                t=40,
                                pad=4),
                     )
    
    fig.add_scatter(x=x_data, y=y_data, name='Данные расчетные', mode='lines', 
                    line=dict(color='#191970', width=3),
                     row=1,col=1)
    fig.add_scatter(x=x_data, y=z_data, name='Данные аналоги', mode='lines', 
                    line=dict(color='red', width=3),
                     row=1,col=1)

    return fig

data_pipes=pd.read_excel('doc/excel_pipes.xlsx')
name_list_pipes=[item.replace('\xa0',' ') for item in list(data_pipes.columns)]
data_pipes.columns=name_list_pipes
for i in data_pipes.columns:
    try:data_pipes[i]=data_pipes[i].astype('float')
    except: ...
data_pipes['id']=data_pipes['Диаметр трубы'].apply(lambda x: format(x, '.2f'))+'_'+\
            data_pipes['Стенка трубы'].apply(lambda x: format(x, '.2f'))+'_'+\
            data_pipes['Диаметр оправки (мм)_MPM'].apply(lambda x: format(x, '.2f'))+'_'+\
            data_pipes['Калибр _ELM'].apply(lambda x: format(x, '.2f'))+'_'+\
            data_pipes['марка стали'].astype('str')
data_pipes=data_pipes[[data_pipes.columns[-1]]+list(data_pipes.columns[:-1])]

list_name_db=list(functions.read_sql_ff('dash_pipes').columns)

data_name=pd.DataFrame({'name_dash':list(data_pipes.columns),'name_bd':list_name_db,'type':list(data_pipes.dtypes)})
data_name.loc[data_name['type']=='float64','type']='numeric'
data_name.loc[data_name['type']=='object','type']='text'

cols0=[{"name": data_name['name_dash'][i], "id": data_name['name_bd'][i], 'type': data_name['type'][i]} for i in range(len(data_name)) ]

def get_nearest_value(iterable, value):
    return min(iterable, key=lambda x: abs(x - value))

Kalibr=0
diam_fact_mpm=0

def ff_speed(zazor,oborot,Kalibr=Kalibr,diam_fact_mpm=diam_fact_mpm,data_prop=data_prop):
    data_speed=data_prop.loc[data_prop['Калибр']==get_nearest_value(data_prop['Калибр'], Kalibr)]
 
    data_speed['Диаметр бочки валка, мм']=diam_fact_mpm
    data_speed['зазор, мм']=zazor
    data_speed['обороты двигателя, об/мин']=oborot
    data_speed['Скорость валков, об/мин']=data_speed['обороты двигателя, об/мин']/data_speed['передаточное число редуктора']
    data_speed['диаметр калибра валка МРМ, мм']=2*data_speed['высота калибра, мм']+data_speed['зазор, мм']
    data_speed['Катающий диаметр, мм']=data_speed['Диаметр бочки валка, мм']+data_speed['зазор, мм']-data_speed['Коэффициент формы калибра']*data_speed['диаметр калибра валка МРМ, мм']
 
    data_speed['скорость трубы в клети, мм/сек']=3.14*data_speed['Катающий диаметр, мм']*data_speed['Скорость валков, об/мин']/60
    fig=make_subplots(rows=1,cols=1)
    fig = px.line(data_speed, x="№ клети", y="скорость трубы в клети, мм/сек")
    fig.update_layout(barmode='relative', 
                      autosize = True,
                      yaxis_title ='Cкорость трубы в клети, мм/сек',
                      xaxis_title ='Номер клети',
                      margin=dict(
                            l=10,
                            r=20,
                            b=100,
                            t=40,
                            pad=4),)
    # fig.show()
    # return list(data_speed['скорость трубы в клети, мм/сек'])
    return fig

def table_1(data,row,selected_rows=[],data_name=data_name):
    if (row!=None) and (row!=[]):
        data2=data.loc[[row[0]]]
        ############### table
        table_1=dash_table.DataTable(
                id='out-table',
                columns=cols0,
                data=data.to_dict('records'),
                page_size=10,
                style_cell_conditional=[{'if': {'column_id': c},'display': 'none'} for c in [list(data_name['name_bd'])[0]]+list(data_name['name_bd'])[6:]],
                # editable=editable,
                style_cell={
                'whiteSpace': 'normal',
                'height': 'auto',
                'overflow': 'hidden',
                'textAlign': 'center',
                },
                row_deletable=False,
                filter_action="native",
                sort_action="native",
                row_selectable="single",
                sort_mode="single",  
                page_action='native',
                selected_rows=selected_rows,
                css=[{'selector': '.dash-spreadsheet tr', 'rule': 'height: 22px'}],
                style_header={
                'textAlign': 'center',
                'font-weight': 'bold',
                 'backgroundColor': "#F5F5F5",},
                page_current= 0)
        ############### table2
        data_mpm_esm=pd.DataFrame({'стан':['МПМ','МПМ','МПМ','МПМ','МПМ','МПМ','МПМ','ИКС','ИКС','ИКС','ИКС','ИКС','ИКС','ИКС','ИКС','ИКС']})
        data_mpm_esm['Номер клети']=[1,2,3,4,5,6,7,1,2,3,4,5,6,7,8,9]
        data_mpm_esm['Диаметр Валков']=list(data2[['diametrvalkovmmklet1mpm','diametrvalkovmmklet2mpm',\
                                                   'diametrvalkovmmklet3mpm','diametrvalkovmmklet4mpm',\
                    'diametrvalkovmmklet5mpm','diametrvalkovmmklet6mpm','diametrvalkovmmklet7mpm',\
                                                   'diametrvalkovmmklet1esm','diametrvalkovmmklet2esm',\
                        'diametrvalkovmmklet3esm','diametrvalkovmmklet4esm','diametrvalkovmmklet5esm',\
                                                   'diametrvalkovmmklet6esm','diametrvalkovmmklet7esm',\
                            'diametrvalkovmmklet8esm','diametrvalkovmmklet9esm']].mean())
        data_mpm_esm['Обороты']=list(data2[['skorostvalkovklet1obminmpm','skorostvalkovklet2obminmpm',\
                                            'skorostvalkovklet3obminmpm',\
            'skorostvalkovklet4obminmpm','skorostvalkovklet5obminmpm','skorostvalkovklet6obminmpm',\
                                            'skorostvalkovklet7obminmpm','skorostkletiobminklet1esm',\
                'skorostkletiobminklet2esm','skorostkletiobminklet3esm','skorostkletiobminklet4esm',\
                                            'skorostkletiobminklet5esm','skorostkletiobminklet6esm',\
                    'skorostkletiobminklet7esm','skorostkletiobminklet8esm','skorostkletiobminklet9esm']].mean())
        data_mpm_esm['Зазоры']=list(data2[['zazorvalkovmmklet1mpm', 'zazorvalkovmmklet2mpm',\
                                           'zazorvalkovmmklet3mpm','zazorvalkovmmklet4mpm',\
            'zazorvalkovmmklet5mpm','zazorvalkovmmklet6mpm', 'zazorvalkovmmklet7mpm',\
                                           'zazorvalkovmmklet1esm', 'zazorvalkovmmklet2esm','zazorvalkovmmklet3esm',\
                'zazorvalkovmmklet4esm', 'zazorvalkovmmklet5esm', 'zazorvalkovmmklet6esm', 'zazorvalkovmmklet7esm',\
                                           'zazorvalkovmmklet8esm', 'zazorvalkovmmklet9esm']].mean())
        ############### plot_1
        zazor=data2[['zazorvalkovmmklet1mpm','zazorvalkovmmklet2mpm','zazorvalkovmmklet3mpm','zazorvalkovmmklet4mpm',\
                     'zazorvalkovmmklet5mpm', 'zazorvalkovmmklet6mpm','zazorvalkovmmklet7mpm']]
        zazor=list(zazor.mean())

        oborot=data2[['skorostvalkovklet1obminmpm','skorostvalkovklet2obminmpm','skorostvalkovklet3obminmpm',\
                      'skorostvalkovklet4obminmpm', 'skorostvalkovklet5obminmpm','skorostvalkovklet6obminmpm','skorostvalkovklet7obminmpm']]
        oborot=list(oborot.mean())

        Kalibr=list(data2['kalibrelm'])[0]

        diam_fact_mpm=data2[['diametrvalkovmmklet1mpm','diametrvalkovmmklet2mpm','diametrvalkovmmklet3mpm',\
                             'diametrvalkovmmklet4mpm',\
            'diametrvalkovmmklet5mpm','diametrvalkovmmklet6mpm','diametrvalkovmmklet7mpm']]
        diam_fact_mpm=list(diam_fact_mpm.mean())
         
        list_a = ["стан", 'Номер клети']

        cols_tb2 = [
                {"name": i, "id": i, "editable": False,} for i in data_mpm_esm.columns if i in list_a]

        cols_tb2_2 = [
                {"name": i, "id": i, 'type': 'numeric', "editable": True,
                 } for i in data_mpm_esm.columns if i not in list_a]

        cols_tb2.extend(cols_tb2_2)

        table_2=dash_table.DataTable(
            id='out-table2',
            #columns=[{"name": i, "id": i} for i in data_mpm_esm.columns],
            columns=cols_tb2,
            data=data_mpm_esm.to_dict('records'),
            css=[{'selector': '.dash-spreadsheet tr', 'rule': 'height: 22px'}],
            style_header={
                'textAlign': 'center',
                'font-weight': 'bold',
                 'backgroundColor': "#F5F5F5",},
            style_cell={
                'whiteSpace': 'normal',
                'height': 'auto',
                'overflow': 'hidden',
                'textAlign': 'center',
                },
            style_cell_conditional=[{'if': {'row_index': 'odd'},
            'backgroundColor': "#F5F5F5",}],
            editable=True)
            
            
        plot_1=dcc.Graph(figure=ff_speed(zazor=zazor,
                    oborot=oborot,
                    Kalibr=Kalibr,
                    diam_fact_mpm=diam_fact_mpm,
                    data_prop=data_prop))
        return table_1,table_2,plot_1
    else:
        table_1=dash_table.DataTable(
                id='out-table',
                columns=cols0,
                data=data.to_dict('records'),
                page_size=10,
                style_cell_conditional=[{'if': {'column_id': c},'display': 'none'} for c in [list(data_name['name_bd'])[0]]+list(data_name['name_bd'])[6:]],
                # editable=editable,
                row_deletable=False,
                filter_action="native",
                sort_action="native",
                row_selectable="single",
                sort_mode="single",  
                page_action='native',
                selected_rows=selected_rows,
                css=[{'selector': '.dash-spreadsheet tr', 'rule': 'height: 22px'}],
                style_header={
                'textAlign': 'center',
                'font-weight': 'bold',
                 'backgroundColor': "#F5F5F5",},
                page_current= 0)
        return table_1,no_update,no_update

table_2=dash_table.DataTable(id='out-table2',data=pd.DataFrame().to_dict('records'), 
                css=[{'selector': '.dash-spreadsheet tr', 'rule': 'height: 22px'}],
                style_header={
                'textAlign': 'center',
                'font-weight': 'bold',
                'backgroundColor': "#F5F5F5",},)

fig=make_subplots(rows=1,cols=1)
fig.update_layout(barmode='relative', title_text='график трубы')
fig.add_scatter(x=[1,2,3,4,5,6,7], y=[830,1300,1600,1800,1870,1900,1860],mode='lines',name="исторические данные",row=1,col=1)
plot_1=dcc.Graph(figure=fig)

def df_to_csv_2(row,dataset,dataset2): 

    pg = pd.DataFrame(dataset)
    pg2 = pd.DataFrame(dataset2)

    pg2['Диаметр Валков']=pg2['Диаметр Валков'].astype('float')
    pg2['Обороты']=pg2['Обороты'].astype('float')
    pg2['Зазоры']=pg2['Зазоры'].astype('float')

    pg.loc[row,['diametrvalkovmmklet1mpm','diametrvalkovmmklet2mpm','diametrvalkovmmklet3mpm','diametrvalkovmmklet4mpm',\
                            'diametrvalkovmmklet5mpm','diametrvalkovmmklet6mpm','diametrvalkovmmklet7mpm','diametrvalkovmmklet1esm','diametrvalkovmmklet2esm',\
                                'diametrvalkovmmklet3esm','diametrvalkovmmklet4esm','diametrvalkovmmklet5esm','diametrvalkovmmklet6esm','diametrvalkovmmklet7esm',\
                                    'diametrvalkovmmklet8esm','diametrvalkovmmklet9esm']]=list(pg2['Диаметр Валков'])
    pg.loc[row,['skorostvalkovklet1obminmpm','skorostvalkovklet2obminmpm','skorostvalkovklet3obminmpm',\
                    'skorostvalkovklet4obminmpm','skorostvalkovklet5obminmpm','skorostvalkovklet6obminmpm','skorostvalkovklet7obminmpm','skorostkletiobminklet1esm',\
                        'skorostkletiobminklet2esm','skorostkletiobminklet3esm','skorostkletiobminklet4esm','skorostkletiobminklet5esm','skorostkletiobminklet6esm',\
                            'skorostkletiobminklet7esm','skorostkletiobminklet8esm','skorostkletiobminklet9esm']]=list(pg2['Обороты'])
    pg.loc[row,['zazorvalkovmmklet1mpm', 'zazorvalkovmmklet2mpm', 'zazorvalkovmmklet3mpm','zazorvalkovmmklet4mpm',\
                    'zazorvalkovmmklet5mpm','zazorvalkovmmklet6mpm', 'zazorvalkovmmklet7mpm','zazorvalkovmmklet1esm', 'zazorvalkovmmklet2esm','zazorvalkovmmklet3esm',\
                        'zazorvalkovmmklet4esm', 'zazorvalkovmmklet5esm', 'zazorvalkovmmklet6esm', 'zazorvalkovmmklet7esm', 'zazorvalkovmmklet8esm',\
                            'zazorvalkovmmklet9esm']]=list(pg2['Зазоры'])

    pg['id']=pg['diametrtruby'].apply(lambda x: format(float(x), '.2f'))+'_'+\
    pg['stenkatruby'].apply(lambda x: format(float(x), '.2f'))+'_'+\
    pg['diametropravkimmmpm'].apply(lambda x: format(float(x), '.2f'))+'_'+\
    pg['kalibrelm'].apply(lambda x: format(float(x), '.2f'))+'_'+\
    pg['markastali'].astype('str')

    pg=pg.replace('', 0)

    for i in pg.columns:
        if i=='markastali' or i=='id':
            pg[i]=pg[i].astype('object')
        else:
            pg[i]=pg[i].astype('float')

    functions.update_sql_ff(pg,'dash_pipes')
    return html.Plaintext("данные сохранены "+str(datetime.now() + timedelta(hours=3)),style={'color': 'red', 'font-weight': 'bold', 'font-size': 'large'})

diam_fact_mpm= [916,894,912,817,818,826,827]
diam_fact_mpm_new= [1016,994,1012,917,918,926,927]
def return_speed(zazor,oborot,Kalibr=Kalibr,diam_fact_mpm=diam_fact_mpm,data_prop=data_prop,diam_fact_mpm_new=diam_fact_mpm_new):
    data_speed=data_prop.loc[data_prop['Калибр']==get_nearest_value(data_prop['Калибр'], Kalibr)]
 
    data_speed['Диаметр бочки валка, мм']=diam_fact_mpm
    data_speed['зазор, мм']=zazor
    data_speed['обороты двигателя, об/мин']=oborot
    data_speed['Скорость валков, об/мин']=data_speed['обороты двигателя, об/мин']/data_speed['передаточное число редуктора']
    data_speed['диаметр калибра валка МРМ, мм']=2*data_speed['высота калибра, мм']+data_speed['зазор, мм']
    data_speed['Катающий диаметр, мм']=data_speed['Диаметр бочки валка, мм']+data_speed['зазор, мм']-data_speed['Коэффициент формы калибра']*data_speed['диаметр калибра валка МРМ, мм']
 
    data_speed['скорость трубы в клети, мм/сек']=3.14*data_speed['Катающий диаметр, мм']*data_speed['Скорость валков, об/мин']/60
    speed=list(data_speed['скорость трубы в клети, мм/сек'])
    ###
    np_array = np.array(speed)
    np_round_ = np.around(np_array, 1)
    speed = list(np_round_)

    data_speed=data_prop.loc[data_prop['Калибр']==get_nearest_value(data_prop['Калибр'], Kalibr)]
    data_speed['Диаметр бочки валка, мм']=diam_fact_mpm_new
    data_speed['зазор, мм']=zazor
    data_speed['диаметр калибра валка МРМ, мм']=2*data_speed['высота калибра, мм']+data_speed['зазор, мм']
    data_speed['Катающий диаметр, мм']=data_speed['Диаметр бочки валка, мм']+data_speed['зазор, мм']\
        -data_speed['Коэффициент формы калибра']*data_speed['диаметр калибра валка МРМ, мм']

    MPM_oborot=list(pd.Series(speed)*60/3.14/data_speed['Катающий диаметр, мм'].reset_index(drop=True)*data_speed['передаточное число редуктора'].reset_index(drop=True))
    np_array = np.array(MPM_oborot)
    np_round_ = np.around(np_array, 0)
    MPM_oborot = list(np_round_)
    
    diametr_cat = list(data_speed['Катающий диаметр, мм'])
    np_array = np.array(diametr_cat)
    np_round_ = np.around(np_array, 1)
    diametr_cat = list(np_round_)
    return speed,MPM_oborot,diametr_cat #скорости, обороты

def create_analog_tabl(df):
    table = dash_table.DataTable(
        id='table_analog',
        columns= [
        #{'name': 'Время создания', 'id': 'mpmin', "editable": False},
        {'name': 'Время создания', 'id': 'Data', "editable": False},
        {'name': 'Марки стали', 'id': 'markastali', "editable": False},
        {'name': 'Источник данных', 'id': 'Источник данных', "editable": False},
        {'name': 'ID', 'id': 'ID', "editable": False},],
        
        data = df.to_dict('record'),
        row_selectable="single",
        page_size=10,
        selected_rows=[],
        filter_action='native', 
        page_action='native',
        style_as_list_view=True,
            style_cell={
            'whiteSpace': 'normal',
            'height': 'auto',
            'font-size': '14px'
            },   
        style_data_conditional=[
        {
             'if': {'row_index': 'odd'},
            'backgroundColor': "#c7e5f3", 
        }],
        style_cell_conditional=[
        {'if': {'column_id': 'ID'},'display': 'none'},
        {'if': {'column_id': 'Data'},'textAlign': 'center'},
        {'if': {'column_id': 'markastali'},'textAlign': 'center'},
        {'if': {'column_id': 'Источник данных'},'textAlign': 'center'},

    ],
     style_header={
        #"backgroundColor": "#353A40", #292D32
        'textAlign': 'center',
        'font-weight': 'bold',
         'backgroundColor': "#F5F5F5",
        },
    css=[{'selector': '.dash-spreadsheet tr', 'rule': 'height: 18px'}],
   # fixed_rows={'headers': True},
    
    )
    return table

#df для таблицы с зазорами и оборотами. Передает стрку из БД (текущии) + зазоры + обороты аналогов
def tables_zazory(df, df2, skor, zazor):
    df1 = df.copy()
    df2 = df2.T
    df2 = df2.reset_index()
    df2.rename(columns = {'index': 'Report', 0: 'Значение'}, inplace = True)
    df2['Report'] = df2.Report.replace({'vtz':'VTZ'}, regex=True)
    df3 = df1.merge(df2, left_on='Report', right_on='Report', how = 'left')
        
    tb = df3[df3['num'].isin(range(19,68))]
    tb = tb.sort_values(by=['num'])

    tb = tb[['name_properties_2', 'Значение']]
    tb = tb.rename(columns = {'name_properties_2': 'Параметр', 'Значение': 'Значение'}, inplace = False)
    tb = tb.reset_index(drop=True)
    
    df_zazor = pd.DataFrame()
    df_zazor['Стан'] = ['МПМ']*7 + ['ИКС']*9
    df_zazor['Номер клети'] = list(range(1, 8)) + list(range(1,10))
    df_zazor['Диаметр Валков'] = tb.loc[tb['Параметр'] =='Актуал. диаметр валка, мм', 'Значение'].reset_index(drop=True)
    df_zazor['Обороты расчетные'] = tb.loc[tb['Параметр'] =='Скорость двигателей, об/мин', 'Значение'].reset_index(drop=True)
    df_zazor['Зазоры расчетные'] = tb.loc[tb['Параметр'] =='Зазор валков, мм', 'Значение'].reset_index(drop=True)
    
    df_zazor['Обороты аналоги'] = skor
    df_zazor['Зазоры аналоги'] = zazor

    df_zazor['Диаметр Валков'] = (pd.to_numeric(df_zazor['Диаметр Валков'], downcast="float", errors='coerce').fillna(0)).astype(float).round(0)
    df_zazor['Обороты расчетные'] = (pd.to_numeric(df_zazor['Обороты расчетные'], downcast="float", errors='coerce').fillna(0)).astype(float).round(1)
    df_zazor['Обороты аналоги'] = (pd.to_numeric(df_zazor['Обороты аналоги'], downcast="float", errors='coerce').fillna(0)).astype(float).round(1)
    df_zazor['Зазоры расчетные'] = (pd.to_numeric(df_zazor['Зазоры расчетные'], downcast="float", errors='coerce').fillna(0)).astype(float).round(1)
    df_zazor['Зазоры аналоги'] = (pd.to_numeric(df_zazor['Зазоры аналоги'], downcast="float", errors='coerce').fillna(0)).astype(float).round(1)
    
    df_zazor.replace([np.inf, -np.inf, np.nan], 0, inplace=True)
    df_zazor['Зазоры расчетные'] = df_zazor.apply(lambda row: 0 if row['Диаметр Валков'] == 0  else row['Зазоры расчетные'], axis=1)
    df_zazor['Обороты расчетные'] = df_zazor.apply(lambda row: 0 if row['Диаметр Валков'] == 0 else row['Обороты расчетные'], axis=1)
    df_zazor['Обороты аналоги'] = df_zazor.apply(lambda row: 0 if row['Диаметр Валков'] == 0 else row['Обороты аналоги'], axis=1)
    df_zazor['Зазоры аналоги'] = df_zazor.apply(lambda row: 0 if row['Диаметр Валков'] == 0 else row['Зазоры аналоги'], axis=1)
    
    df_zazor['Разница зазоры']  = (df_zazor['Зазоры расчетные'] - df_zazor['Зазоры аналоги']).astype(float).round(1)
    df_zazor['Разница обороты,%']  = ((df_zazor['Обороты расчетные'] - df_zazor['Обороты аналоги'])/df_zazor['Обороты расчетные'] *100).astype(float).round(1)

    df_zazor.replace([np.inf, -np.inf, np.nan], 0, inplace=True)

    return df_zazor

def table_realpipes(real_pipes, row):
    try:
        real_pipes = real_pipes.rename(columns={'timestamp': 'Дата обкатки',
                                        'diametropravkimmmpm': 'Диаметр оправки (мм)_МРМ', 
                                        'diametrtruby': 'Диаметр трубы',
                                        'stenkatruby': 'Стенка трубы',
                                       'markastali': 'Марка стали',
                                       'kalibrelm': 'Калибр_ELM'})
        cols = list(real_pipes.columns)
        cols[3], cols[4] = cols[4], cols[3]
        real_pipes = real_pipes[cols]
            
    except:
        real_pipes = pd.DataFrame(columns=['Дата обкатки', 'Диаметр оправки (мм)_МРМ',
                                          'Диаметр трубы', 'Стенка трубы', 'Марка стали', 'Калибр_ELM'])
        
    table_1= dash_table.DataTable(
            id='real_pipes',
            columns= [{"name": i, "id": i} 
                     for i in real_pipes.columns],
            data=real_pipes.to_dict('records'),
            page_size=10,
            style_cell_conditional=[{'if': {'column_id': c},'display': 'none'} 
                                        for c in list(real_pipes.columns)[6:]],
            # editable=editable,
            row_deletable=False,
            filter_action="native",
            sort_action="native",
            row_selectable="single",
            sort_mode="single",  
            page_action='native',
            css=[{'selector': '.dash-spreadsheet tr', 'rule': 'height: 22px'}],
            style_cell={
                'whiteSpace': 'normal',
                'height': 'auto',
                'overflow': 'hidden',
                'textAlign': 'center',
                # 'font-size': '12px'
                },
           
            style_header={
                'textAlign': 'center',
                'font-weight': 'bold',
                 'backgroundColor': "#F5F5F5",},
            page_current= 0),
    
     ############### table2
    if len(real_pipes) >0:
        data2=real_pipes.loc[row]
        Kalibr=data2['Калибр_ELM']

        data_mpm_esm=pd.DataFrame({'стан':['МПМ','МПМ','МПМ','МПМ','МПМ','МПМ','МПМ','ИКС','ИКС','ИКС','ИКС','ИКС','ИКС','ИКС','ИКС','ИКС']})
        data_mpm_esm['Номер клети']=[1,2,3,4,5,6,7,1,2,3,4,5,6,7,8,9]
        data_mpm_esm['Диаметр Валков']=list(data2[['diametrvalkovmmklet1mpm','diametrvalkovmmklet2mpm','diametrvalkovmmklet3mpm','diametrvalkovmmklet4mpm',\
                        'diametrvalkovmmklet5mpm','diametrvalkovmmklet6mpm','diametrvalkovmmklet7mpm','diametrvalkovmmklet1esm','diametrvalkovmmklet2esm',\
                            'diametrvalkovmmklet3esm','diametrvalkovmmklet4esm','diametrvalkovmmklet5esm','diametrvalkovmmklet6esm','diametrvalkovmmklet7esm',\
                                'diametrvalkovmmklet8esm','diametrvalkovmmklet9esm']])
        data_mpm_esm['Обороты']=list(data2[['skorostvalkovklet1obminmpm','skorostvalkovklet2obminmpm','skorostvalkovklet3obminmpm',\
                'skorostvalkovklet4obminmpm','skorostvalkovklet5obminmpm','skorostvalkovklet6obminmpm','skorostvalkovklet7obminmpm','skorostkletiobminklet1esm',\
                    'skorostkletiobminklet2esm','skorostkletiobminklet3esm','skorostkletiobminklet4esm','skorostkletiobminklet5esm','skorostkletiobminklet6esm',\
                        'skorostkletiobminklet7esm','skorostkletiobminklet8esm','skorostkletiobminklet9esm']])
        data_mpm_esm['Зазоры']=list(data2[['zazorvalkovmmklet1mpm', 'zazorvalkovmmklet2mpm', 'zazorvalkovmmklet3mpm','zazorvalkovmmklet4mpm',\
                'zazorvalkovmmklet5mpm','zazorvalkovmmklet6mpm', 'zazorvalkovmmklet7mpm','zazorvalkovmmklet1esm', 'zazorvalkovmmklet2esm','zazorvalkovmmklet3esm',\
                    'zazorvalkovmmklet4esm', 'zazorvalkovmmklet5esm', 'zazorvalkovmmklet6esm', 'zazorvalkovmmklet7esm', 'zazorvalkovmmklet8esm',\
                        'zazorvalkovmmklet9esm']])

        #         ############### plot_1
        zazor=data2[['zazorvalkovmmklet1mpm','zazorvalkovmmklet2mpm','zazorvalkovmmklet3mpm','zazorvalkovmmklet4mpm','zazorvalkovmmklet5mpm',\
                'zazorvalkovmmklet6mpm','zazorvalkovmmklet7mpm']]

        oborot=data2[['skorostvalkovklet1obminmpm','skorostvalkovklet2obminmpm','skorostvalkovklet3obminmpm','skorostvalkovklet4obminmpm',\
                'skorostvalkovklet5obminmpm','skorostvalkovklet6obminmpm','skorostvalkovklet7obminmpm']]

        diam_fact_mpm=data2[['diametrvalkovmmklet1mpm','diametrvalkovmmklet2mpm','diametrvalkovmmklet3mpm','diametrvalkovmmklet4mpm',\
            'diametrvalkovmmklet5mpm','diametrvalkovmmklet6mpm','diametrvalkovmmklet7mpm']]
        
        plot_1=dcc.Graph(figure=ff_speed(zazor=list(zazor.astype(float)),
                    oborot=list(oborot.astype(float)),
                    Kalibr=float(Kalibr),
                    diam_fact_mpm=list(diam_fact_mpm.astype(float)),
                    data_prop=data_prop))
       
    else:
        plot_1= None
        data_mpm_esm = pd.DataFrame()
        
    list_a = ["стан", 'Номер клети']
    cols_tb2 = [{"name": i, "id": i, "editable": False,} for i in data_mpm_esm.columns if i in list_a]
    cols_tb2_2 = [{"name": i, "id": i, 'type': 'numeric', "editable": True,} for i in data_mpm_esm.columns if i not in list_a]
    cols_tb2.extend(cols_tb2_2)

    table_2=dash_table.DataTable(
                id='real_pipes2',
                #columns=[{"name": i, "id": i} for i in data_mpm_esm.columns],
                columns=cols_tb2,
                data=data_mpm_esm.to_dict('records'),
                css=[{'selector': '.dash-spreadsheet tr', 'rule': 'height: 22px'}],
                style_header={
                    'textAlign': 'center',
                    'font-weight': 'bold',
                     'backgroundColor': "#F5F5F5",},
                style_cell={
                    'whiteSpace': 'normal',
                    'height': 'auto',
                    'overflow': 'hidden',
                    'textAlign': 'center',
                    # 'font-size': '12px'
                    },
                 style_cell_conditional=[{'if': {'row_index': 'odd'},
                'backgroundColor': "#F5F5F5",}],
                editable=True)
            

    return table_1,table_2,plot_1
 
def currents_graph(currents):
    #change data types to float
    for col in range(1,(currents.shape[1])):
        name = currents.columns[col]
        try:
            currents[name] = currents[name].astype("float")
        except:
            currents[name] = currents[name].astype("object")
    

    try:
        fig=make_subplots(rows=1,cols=1,specs=[[{"secondary_y": True}]])

        for i in range(1,8):
            fig.add_scatter(x=currents['timestamp'], y=currents['tas_mpm_curr_std'+str(i)], name='ток'+str(i), mode='lines',row=1,col=1,secondary_y=False)

        fig.update_layout(barmode='relative',     
                              title={
                'text': "Токи mpm",
                'y':0.99,
                'x':0.5,
                'xanchor': 'center',
                'yanchor': 'top'},
            font=dict(
        #         family="Courier New, monospace",
        #         size=18,
                color="grey"),
            legend=dict(
#                                     orientation="h",
#                                     yanchor="bottom",
#                                     y=1.01,
#                                     xanchor="left",
#                                     x=0.0,
                        font=dict(size=10,),
                                ),
                autosize = True,
                           margin=dict(
                                l=5,
                                r=5,
                                b=5,
                                t=40,
                                pad=4),
            )

        toki_mpm = fig
    except:
        toki_mpm = None

    try:
        fig=make_subplots(rows=1,cols=1,specs=[[{"secondary_y": True}]])

        for i in range(1,10):
            fig.add_scatter(x=currents['timestamp'], y=currents['tas_esm_curr_std'+str(i)], name='ток'+str(i),mode='lines',row=1,col=1,secondary_y=False)

        fig.update_layout(barmode='relative',     
                              title={
                'text': "Токи esm",
                'y':0.99,
                'x':0.5,
                'xanchor': 'center',
                'yanchor': 'top'},
            font=dict(
        #         family="Courier New, monospace",
        #         size=18,
                color="grey"),
            legend=dict(
#                                     orientation="h",
#                                     yanchor="bottom",
#                                     y=1.01,
#                                     xanchor="left",
#                                     x=0.0,
                font=dict(size=10,),
                                ),
            autosize = True,
                           margin=dict(
                                l=5,
                                r=5,
                                b=5,
                                t=40,
                                pad=4),
            )

        toki_esm = fig
    except:
        toki_esm = None
    
    if len(currents) > 0:
        if ff(currents.loc[:,'tas_mpm_curr_std1':'tas_mpm_curr_std7'].T.isnull().sum()) == 7:
            toki_mpm = None
        if ff(currents.loc[:,'tas_esm_curr_std1':'tas_esm_curr_std9'].T.isnull().sum()) == 9:
            toki_esm = None
    elif len(currents) == 0:
        toki_mpm = None
        toki_esm = None

    return toki_mpm, toki_esm

#UPDATE VERSION функция возвращает таблицу с аналогами по условиям 
def slice_tb(x, id_):

    sql_query1 = 'WHERE diametrtruby = (%s) AND stenkatruby = (%s) AND  kalibrelm = (%s) AND  diametropravkimmmpm = (%s) ORDER BY mpmin DESC LIMIT 100'
    new_tbl = functions2.data_query(sql_query1, [float(i) for i in ff(x)],  data = 'dash_analogue')
    new_tbl['Источник данных'] = 'Данные со стана'

    new_tbl = new_tbl[['mpmin', 'markastali', 'Источник данных', 'idpipesinuque']]
    new_tbl.rename(columns = {"mpmin": "Data", "idpipesinuque": "ID"}, 
          inplace = True)

    sql_query2 = 'WHERE diametrtruby = (%s) AND stenkatruby = (%s) AND  kalibrelm = (%s) AND  diametropravkimmmpm = (%s) ORDER BY ctid DESC LIMIT 100'
    new_tbl2 = functions2.data_query(sql_query2, [float(i) for i in ff(x)],  data = 'dash_pipes')
    new_tbl2['Источник данных'] = 'Фактические настройки'

    new_tbl2['Data'] = None
    new_tbl2 = new_tbl2[['Data', 'markastali', 'Источник данных', 'id']]
    new_tbl2.rename(columns = {"id": "ID"}, 
          inplace = True)
    
    sql_query3 = 'WHERE vtz_report2_1 = (%s) AND vtz_report2_3 = (%s) AND  vtz_report1_1 = (%s) AND vtz_report1_2 = (%s) ORDER BY datetime DESC LIMIT 100'
    
    new_tbl3 = functions2.data_query(sql_query3, ff(x), data = 'data_report')
    new_tbl3 = new_tbl3[(new_tbl3['scenarioid'] != id_)]
    new_tbl3['Источник данных'] = 'Предварительные расчеты'
    new_tbl3.rename(columns = {"datetime": "Data", "scenarioid": "ID", 'vtz_report1_3': 'markastali'}, 
          inplace = True)
    new_tbl3 = new_tbl3[['Data', 'markastali', 'Источник данных', 'ID']]
    
    
    sql_query4 = 'WHERE diametrtruby = (%s) AND stenkatruby = (%s) AND  kalibrelm = (%s) AND  diametropravkimmmpm = (%s) ORDER BY timestamp DESC LIMIT 100'
    new_tbl4 = functions2.data_query(sql_query4, [float(i) for i in ff(x)],  data = 'integrazaya')
    new_tbl4['Источник данных'] = 'Данные со стана (интеграция)'

    new_tbl4 = new_tbl4[['timestamp', 'markastali', 'Источник данных', 'id']]
    new_tbl4.rename(columns = {"timestamp": "Data", "id": "ID"}, inplace = True)
    
    Table = pd.concat([new_tbl, new_tbl2, new_tbl3, new_tbl4])
    Table['Data'] = Table['Data'].astype('datetime64[ns]').dt.round('S').dt.strftime('%Y-%m-%d %H:%M:%S')

    return Table

#app = dash.Dash(__name__, external_stylesheets=[dbc.themes.FLATLY], prevent_initial_callbacks=True)
app = dash.Dash(__name__, prevent_initial_callbacks=True)
server = app.server


#ТАБЛИЦЫ ДЛЯ ВВОДА
TABLE_01 = dash_table.DataTable(
        id='table-01',
        columns= [
        {'name': 'Параметр', 'id': 'Параметр', "editable": False},
        {'name': 'Значение', 'id': 'Значение', "editable": True},],

        data = display_tb(df, df2 = pd.DataFrame())[0].to_dict('records'),
        cell_selectable = True,
        #selected_columns=[],
        style_as_list_view=True,
            style_cell={
            'whiteSpace': 'normal',
            #'height': 'auto',
            #'height': '13px',
            #'font-size': '12px',
            #'minWidth': '150%', 'width': '150%', 'maxWidth': '150%',
              #  'height': 5,
            #'overflow': 'hidden',
            #'textOverflow': 'ellipsis',
           # 'maxWidth': 0,
           # "backgroundColor": "#58606A",  
           # 'color': 'white',
            },   
    
        style_data_conditional=[
        {
             'if': {'row_index': 4},
            'backgroundColor': "#F5F5F5",
            'font-weight': 'bold',
        }],

       dropdown_conditional=[{
            'if': {
                'column_id': 'Значение',
                #'row_index': 2,
                'filter_query': '{Параметр} eq "Марка стали"'
            },
            'options': [
                            {'label': i, 'value': i}
                            for i in list_marki_1
                        ]
        },
       {
            'if': {
                'column_id': 'Значение',
                #'row_index': 2,
                'filter_query': '{Параметр} eq "Калибр, мм"'
            },
            'options': [
                            {'label': i, 'value': i}
                            for i in list_kalibr_1
                        ]
        }],
        #editable=[{'Значение': True, 'Параметр': False}],
        style_cell_conditional=[
#         {
#             'if': {'column_id': c},
#             'width': '150%',
#             'textAlign': 'left'
#         } for c in ['Параметр', 'Значение']
            
        {'if': {'column_id': 'Параметр'},
         'width': '70%', 'textAlign': 'left'},
        {'if': {'column_id': 'Значение'},
         'width': '30%', 'textAlign': 'left'},
    ],
     style_header={
        #"backgroundColor": "#353A40", #292D32
        'textAlign': 'left',
        'font-weight': 'bold',
         'backgroundColor': "#F5F5F5",
        },
    css=[{'selector': '.dash-spreadsheet tr', 'rule': 'height: 20px'},
        {"selector": ".Select-menu-outer", "rule": "display: block !important"},],
    )



def table2(df):
    TABLE_02 = dash_table.DataTable(
            id='table-02',
            columns= [
            {'name': 'Параметр', 'id': 'Параметр', "editable": False},
            {'name': 'Значение', 'id': 'Значение', "editable": True, 'type': 'numeric'},],
            data = df.to_dict('records'),
            cell_selectable = True,
            #selected_columns=[],
            style_as_list_view=True,
                style_cell={
                'whiteSpace': 'normal',
                'height': 'auto',
                #'font-size': '12px'
                },   

            style_data_conditional=[
            {'if': {'row_index': c},
                'backgroundColor': "#c7e5f3",
                #'font-weight': 'bold',
            } for c in [5,6,7,8,9,10]],
            #editable=[{'Значение': True, 'Параметр': False}],
            style_cell_conditional=[
#             {
#                 'if': {'column_id': c},
#                 #'width': '150%',
#                 'textAlign': 'left'
#             } for c in ['Параметр', 'Значение'],
        {'if': {'column_id': 'Параметр'},
         'width': '70%', 'textAlign': 'left'},
        {'if': {'column_id': 'Значение'},
         'width': '30%', 'textAlign': 'left'},
        ],
         style_header={
            #"backgroundColor": "#353A40", #292D32
            'textAlign': 'left',
            'font-weight': 'bold',
             'backgroundColor': "#F5F5F5",
            },
        css=[{'selector': '.dash-spreadsheet tr', 'rule': 'height: 20px'}],
        )

    return TABLE_02


TABLE_03 = dash_table.DataTable(
        id='table-03',
        columns= [
        {'name': 'Параметр', 'id': 'Параметр', "editable": False},
        {'name': 'Значение', 'id': 'Значение', "editable": True, 'type': 'numeric'},],
        data = display_tb(df, df2 = pd.DataFrame())[2].iloc[[0,1, 4,5, 8,9,12,13, 16, 17, 20,21,24,25, 28]].to_dict('records'),
        cell_selectable = True,
        #selected_columns=[],
        style_as_list_view=True,
            style_cell={
            'whiteSpace': 'normal',
            'height': 'auto',
           # 'font-size': '12px'
            },   

        style_data_conditional=(
        [{'if': {'row_index': d},
            'backgroundColor': "#F5F5F5",
            'font-weight': 'bold',
        } for d in [0, 2, 4, 6, 8, 10, 12]
    ]),

    
    style_cell_conditional=([
        {
            'if': {'column_id': c},
            'textAlign': 'left'
        } for c in ['Параметр', 'Значение']] 
    ),
     style_header={
        #"backgroundColor": "#353A40", #292D32
        'textAlign': 'left',
        'font-weight': 'bold',
         'backgroundColor': "#F5F5F5",
        },
    css=[{'selector': '.dash-spreadsheet tr', 'rule': 'height: 20px'}],
    #virtualization=True,
    )


TABLE_04 = dash_table.DataTable(
        id='table-04',
        columns= [
        {'name': 'Параметр', 'id': 'Параметр', "editable": False},
        {'name': 'Значение', 'id': 'Значение', "editable": True, 'type': 'numeric'},],
        data = display_tb(df, df2 = pd.DataFrame())[3].iloc[[0,1,4,5, 8,9,12,13, 16, 17, 20,21,24,25, 28, 29, 32, 33]].to_dict('record'),
        cell_selectable = True,
        #selected_columns=[],
        style_as_list_view=True,
            style_cell={
            'whiteSpace': 'normal',
            'height': 'auto',
           # 'font-size': '12px'
            },   
    
        style_data_conditional=(
        [{'if': {'row_index': d},
            'backgroundColor': "#F5F5F5",
            'font-weight': 'bold',
        } for d in [0, 2, 4, 6, 8, 10, 12, 14, 16]
        ]),
    
    style_cell_conditional=[
        {
            'if': {'column_id': c},
            'textAlign': 'left'
        } for c in ['Параметр', 'Значение']
    ],
     style_header={
        #"backgroundColor": "#353A40", #292D32
        'textAlign': 'left',
        'font-weight': 'bold',
         'backgroundColor': "#F5F5F5",
        },
    css=[{'selector': '.dash-spreadsheet tr', 'rule': 'height: 20px'}],
   # fixed_rows={'headers': True},
    
    )


pg1 = choose_table_create(functions2.sql_data_LIMIT500(limit = 3))
col_names = ['Сценарий', 'Дата создания', 'Калибр, мм', 'Код оправки', 
              'Марка стали', 'Диаметр трубы, мм', 'Толщина стенки, мм' ]

#функия для создания таблицы с выбором аналогов
def table_modal(pg, col_names):
    chose_table = dash_table.DataTable(
            id='select_table',
            columns= [{"name": j, "id": i} 
                     for i,j in zip(pg.columns, col_names)],

            data = pg.to_dict('record'),
            row_selectable = 'single',
            filter_action="native",
            #page_action='native',
            page_current= 0,
            page_size= 10,
            style_cell={
                'whiteSpace': 'normal',
                'height': 'auto',
                'overflow': 'hidden',
                 'font-size': '12px'},
        style_cell_conditional=[{'if': {'column_id': 'scenarioid'},'display': 'none'}],
    css=[{'selector': '.dash-spreadsheet tr', 'rule': 'height: 15px'}])
    return chose_table

#функия для создания таблицы с зазорами и оборотами
def zazor_table(df):
    zazor_table = dash_table.DataTable(
            id='zazor_table',
            columns= [{"name": i, "id": i} 
                     for i in df.columns],

            data = df.to_dict('record'),
            style_cell={
                'whiteSpace': 'normal',
                'height': 'auto',
                'overflow': 'hidden',
                'textAlign': 'center',
                 'font-size': '12px'},
        style_cell_conditional=[{'if': {'row_index': 'odd'},
            'backgroundColor': "#F5F5F5",},
        ],
        style_header={
                'textAlign': 'center',
                'font-weight': 'bold',
                 'backgroundColor': "#F5F5F5",},
    css=[{'selector': '.dash-spreadsheet tr', 'rule': 'height: 14px'}])
    return zazor_table


toast_01 = dbc.Toast(
    [TABLE_01],
    header="1. ВВОДНЫЕ ДАННЫЕ", id = 'toast01', style={'margin-bottom': '5px'}
)

toast_02 = dbc.Toast(
    #children = table2(tb2.iloc[0:5]),
    children = table2(display_tb(df, df2 = pd.DataFrame())[1].iloc[0:5]),
    #[TABLE_02],
     #[table2(tb2)],
    header="2. ПРОШИВНОЙ СТАН", id = 'toast02', #style={"maxWidth": "900px"}
)

toast_03 = dbc.Toast(
    [TABLE_03],
    header="3. НЕПРЕРЫВНЫЙ СТАН", id = 'toast03', #style={"maxWidth": "900px"}
)

toast_04 = dbc.Toast(
    [TABLE_04],
    header="4. ИЗВЛЕКАТЕЛЬНО-КАЛИБРОВОЧНЫЙ СТАН", id = 'toast04', #style={"maxWidth": "900px"}
)

style_1={
    'margin-right': '2px', 
}

#### APP #####

app.layout = html.Div([
    dcc.Location(id='url', refresh=False),
    html.Div(id='page-content')
])

#### PAGE 1 #####
####PAGE 1 BUTTONS#####

load_report = html.Div([dbc.DropdownMenu(
            [dbc.DropdownMenuItem("Отчет МПМ", id="item-MPM"), dbc.DropdownMenuItem("Отчет ПС", id="item-PS")],
            label="Технологическая карта", color="primary", bs_size="sm",
            group=True, id = 'dropdown1'),
            Download(id="download"),])
    
report= html.Div([dbc.Button(
    "Расчет", color="primary", size="sm", id = 'model'),
      dbc.Modal(
                            [
                                dbc.ModalHeader("ПРЕДУПРЕЖДЕНИЕ!"),
                                dbc.ModalBody([html.Div("Для запуска расчета, пожалуйста, заполните все данные!"),
                                               html.Br(), 
                                               html.Div("Если ВСЕ данные заполнены, то, пожалуйста, еще раз проверьте введенные параметры (толщину стенки, оправку МПМ, калибр и диаметр трубы, наличие резов)")]),
                                dbc.ModalFooter(
                                    dbc.Button("Закрыть", id="close", className="ml-auto")
                                ),
                            ], 
                            style={"margin-right": '0px !important', 'padding-right': '0px !important'},

                            id="modal",  
                        ),])


load= html.Div([dbc.Button(
    "Выбрать расчет", color="primary", size="sm", id = 'load'),
                     dbc.Modal(
                            [
                                dbc.ModalHeader("Выберите расчет, который Вы хотите загрузить"),
                                dbc.ModalBody(table_modal(pg1, col_names)),
                                #dbc.ModalBody('oppp'),
                                dbc.ModalFooter([
                                    dbc.Button("Загрузить", id="ok", className="ml-auto", color="primary"),
                                    dbc.Button("Удалить", id="del", className="ml-auto", color="primary"),
                                    dbc.Button("Отмена", id="return", className="ml-auto", color="primary"),]
                                ),
                            ], 
                            style={"margin-right": '0px !important', 'padding-right': '0px !important'},

                            id="modal2", size="xl",)
               ])

garbage= dbc.Button(
    "Очистить данные", color="primary", size="sm", id = 'clear')

choose_analog_btn= dbc.Button(
    "Выбрать аналог", color="primary", size="sm", id = 'choose_analog_btn', style={'display': 'none'})

page= dbc.Button(
    "Идеальные трубы", color="primary", size="sm", id = 'page', href='/table')

page_3= dbc.Button(
    "Реальные трубы", color="primary", size="sm", id = 'page_3', href='/realpipes') #переход на страницу 3

group1 = dbc.ButtonGroup(
    [load, report, garbage, load_report, choose_analog_btn,  page, page_3
    ], id = 'buttongrp_1')


# неизменные данные для графика (ocь X) - номера клети
x_data = [1,2,3,4,5,6,7]

###SRYLE FOR UP-D TABLES
style_table1_2=[
        {
             'if': {'row_index': 5},
            'backgroundColor': "#F5F5F5",
            'font-weight': 'bold',
        }]

style_table3=([
        {'if': {'row_index': c},
            'backgroundColor': "#c7e5f3",
        # 'display': 'none',
         #'visibility': 'hidden'
            #'font-weight': 'bold',
        } for c in [2,3,6,7,10,11,14,15,18,19,22,23, 26, 27]] +
        [{'if': {'row_index': d},
            'backgroundColor': "#F5F5F5",
            'font-weight': 'bold',
        } for d in [0, 4, 8, 12, 16, 20, 24]
    ])

style_table4=([
        {'if': {'row_index': c},
            'backgroundColor': "#c7e5f3",
            #'font-weight': 'bold',
        } for c in [2,3,6,7,10,11,14,15,18,19,22,23, 26, 27, 30, 31, 34, 35]] +
        [{'if': {'row_index': d},
            'backgroundColor': "#F5F5F5",
            'font-weight': 'bold',
        } for d in [0, 4, 8, 12, 16, 20, 24, 28, 32]
    ])

style_table1_1=[
        {
             'if': {'row_index': 4},
            'backgroundColor': "#F5F5F5",
            'font-weight': 'bold',
        }]


#### page 1 load analog button ####
save_ideal_btn1 = html.Div([dbc.Button('Сохранить аналог в качестве идеальной трубы', id='save_ideal_btn_pg1', n_clicks=0, size="sm", color="primary",),
                              dbc.Modal(
                            [
                                dbc.ModalHeader(),
                                dbc.ModalBody([
                                    html.P(id = 'load_resultpg1_1'),
                                    ]),
                                dbc.ModalFooter([
                                    dbc.Button("Подтвердить сохранение", id="oksave_page1", className="ml-auto"),
                                    dbc.Button("Отмена", id="rejectsave_page1", className="ml-auto"),]
                                ),
                            ], 
                            style={"margin-right": '0px !important', 'padding-right': '0px !important'},

                            id="modal_analog2_page1", #size="md",
                              ),
                          
                                                        dbc.Modal(
                            [
                                dbc.ModalHeader(),
                                dbc.ModalBody([
                                    html.P(id = 'load_resultpg1_2'),
                                    ]),
                                dbc.ModalFooter([
                                    dbc.Button("Закрыть", id="closeanalogmodal_page1", className="ml-auto"),]
                                ),
                            ], 
                            style={"margin-right": '0px !important', 'padding-right': '0px !important'},

                            id="modal_analog_page1", #size="xl",
                                                        ),
                          
                          ], style= {'display':'none'}, id = 'save_analogpipes_div')
#### PAGE 1 LAYOUT#####
page_1_layout =  dbc.Container([
            
    
            dbc.Row(html.P()),
            group1,

            dbc.Row(html.P()),
            dbc.Row([
                
                
                dbc.Col(html.Div([toast_01, toast_02]), style = style_1,), 
            
                dbc.Col(toast_03, style = style_1,), 

                dbc.Col(toast_04),

            ],
            no_gutters = True), 
           html.P(),  
    
            dbc.Row([dbc.Col([
                html.Div(
                dbc.Card(dcc.Graph(#figure = graph_1(x_data, y_data), 
                                      style = {'height': "22vw"}, 
                                                         id = 'graph_1'
                                     ), body = True, color="light"), id ='hide', style={'display': 'none'}),
                 html.P(),
                html.Div(dash_table.DataTable(id='table_analog'), id = 'table_from_sql',),
                html.P(),
                save_ideal_btn1
            ], width = 6),
                   dbc.Col([html.Div(id = 'div_analog_table', style={'height': '100%'}),
                    ], width = 6,)
                   ]),
            
            html.P(), 
            dcc.Store(id = 'store'),
            dcc.Store(id = 'store2'),
            html.Div(id = 'non_existent', children =
                     dcc.Loading(type='graph', children=html.Div(
            id='loading-hidden-div', children=None, style={'display': 'none'}), 
                        id = 'loading_state', fullscreen = False)
                    ), 
            
                     
            ], fluid = True)


#### PAGE 2 #####

#### PAGE 2 BUTTONS #####
page_2= dbc.Button("Рассчитать зазоры и скорости", color="primary", size="sm", id = 'page_2', href='/') #переход на страницу
add_2 = dbc.Button('Добавить', id='editing-rows-button', n_clicks=0, size="sm", color="primary",)
# add_2 = dcc.Loading(id="loading-2page_1",
#                     children=[dbc.Button('Добавить', id='editing-rows-button', n_clicks=0, size="sm", color="primary",)],
#                     type="circle",)
#save_2 = dbc.Button('Сохранить', id='save_to_postgres', n_clicks=0, size="sm", color="primary",)
save_2 = dcc.Loading(id="loading-2page_2",
                    children=[dbc.Button('Сохранить', id='save_to_postgres', n_clicks=0, size="sm", color="primary",)],
                    type="circle",)

del_2 =  dbc.Button('Удалить', id='del_to_postgres', n_clicks=0, size="sm", color="primary",)

# del_2 =  dcc.Loading(id="loading-2page_3",
#                     children=[dbc.Button('Удалить', id='del_to_postgres', n_clicks=0, size="sm", color="primary",)],
#                     type="circle",)



group2 =  dbc.ButtonGroup([add_2, save_2, del_2, page_2, page_3], id = 'buttongrp_2')
                        
#PAGE 2 LAYOUT#####
page_2_layout = dbc.Container([
    dcc.Interval(id='interval_pg',  interval=86400000*7,n_intervals=0),# 
    html.P(),
    group2,
    html.P(),
    dbc.Row(
            [dbc.Col(html.Div(id='postgres_datatable',
                      children=table_1(data=pd.DataFrame(),
                                       row=None,data_name=data_name)[0],style={'display': 'none'}),width = 12)]),
    html.Button('index', id='index', n_clicks=-1,style={'display': 'none'}),
    html.Div(id='placeholder',children=[]),
    dcc.Interval(id='interval', interval=1000000),
    html.P(),
    dbc.Row(
            [
                dbc.Col(html.Div(id='postgres_datatable2',children=table_2,style={'display': 'none'}), width = 5),
                dbc.Col(html.Div(id='out-plot',children=plot_1,style={'display': 'none'}), width = 7),
            ]
        ),],fluid = True)

#### PAGE 3 #####
#### PAGE 3 BUTTONS #####
kalibrelm = markastali = diametropravkimmmpm = diametrtruby = stenkatruby = []
        
renew = dbc.Button('Обновить данные', id='renew_data', n_clicks=0, size="sm", color="primary",)
# renew = dbc.Button('Обновить', id='renew_data', n_clicks=0, size="sm", color="primary",)
condition_load = html.Div([dbc.Button('Выгрузить данные по условию', id='condition_load_btn', n_clicks=0, size="sm", color="primary",),
                              dbc.Modal(
                            [
                                dbc.ModalHeader("Выберите данные, которые Вы хотите отобразить"),
                                dbc.ModalBody([
                                    html.P(),
                                    dbc.Row([
                                        dbc.Col([
                                            html.Div([html.Div("Калибр_ELM"),
                                            dcc.Dropdown(
                                            options=[{'label': i, 'value': i} for i in kalibrelm],
                                            value= None,
                                            multi=True, id = 'selector_kalibrelm')  
                                              ]),
                                     html.P(),
                                        html.Div([html.Div("Марка стали"),
                                            dcc.Dropdown(
                                            options=[{'label': i, 'value': i} for i in markastali],
                                            value= None,
                                            multi=True, id = 'selector_markastali')  
                                              ]),
                                      html.P(),      
                                 dbc.Row([html.Div('Период времени, за который будут выгружены данные'),
                                        dbc.Col([ #html.Div('C периода'),
                                                     dcc.Input(type= 'date', id = 'date_from')
                                                          ]),
                                        dbc.Col([ #html.Div('По период'),
                                                     dcc.Input(type= 'date', id = 'date_to')]), 
                                            ]),
                                            
                                            ]),
                                    dbc.Col([
                                            html.Div([html.Div("Стенка трубы"),
                                            dcc.Dropdown(
                                            options=[{'label': i, 'value': i} for i in stenkatruby],
                                            value= None,
                                            multi=True, id = 'selector_stenkatruby')  
                                              ]),
                                     html.P(),
                                        html.Div([html.Div("Диаметр трубы"),
                                            dcc.Dropdown(
                                            options=[{'label': i, 'value': i} for i in diametrtruby],
                                            value= None,
                                            multi=True, id = 'selector_diametrtruby')  
                                              ]),
                                    html.P(),
                                        html.Div([html.Div("Диаметр оправки(мм) MPM"),
                                            dcc.Dropdown(
                                            options=[{'label': i, 'value': i} for i in diametropravkimmmpm],
                                            value= None,
                                            multi=True, id = 'selector_diametropravkimmmpm')  
                                              ]),
                                            ]),    
                                html.P(),
                                html.P("Одновременно загружается не более 500 последних по дате труб, подходящих под выбранные условия!", 
                                       style={'color': 'red', 'font-weight': 'bold'})
                                            ])
                                    
                                              ]),
                                dbc.ModalFooter([
                                    dbc.Button("Загрузить", id="ok_page3", className="ml-auto"),
                                    dbc.Button("Отмена", id="return_page3", className="ml-auto"),]
                                ),
                            ], 
                            style={"margin-right": '0px !important', 'padding-right': '0px !important'},

                            id="modal_page3", size="xl",)])

save_ideal_btn = html.Div([dbc.Button('Сохранить как идеальную трубу', id='save_ideal_btn', n_clicks=0, size="sm", color="primary",),
                              dbc.Modal(
                            [
                                dbc.ModalHeader(),
                                dbc.ModalBody([
                                    html.P(id = 'load_result'),
                                    ]),
                                dbc.ModalFooter([
                                    dbc.Button("Подтвердить сохранение", id="oksave_page3", className="ml-auto"),
                                    dbc.Button("Отмена", id="rejectsave_page3", className="ml-auto"),]
                                ),
                            ], 
                            style={"margin-right": '0px !important', 'padding-right': '0px !important'},

                            id="modal_analog2_page3", #size="md",
                              ),
                          
                                                        dbc.Modal(
                            [
                                dbc.ModalHeader(),
                                dbc.ModalBody([
                                    html.P(id = 'load_result2'),
                                    ]),
                                dbc.ModalFooter([
                                    dbc.Button("Закрыть", id="closeanalogmodal_page3", className="ml-auto"),]
                                ),
                            ], 
                            style={"margin-right": '0px !important', 'padding-right': '0px !important'},

                            id="modal_analog_page3", #size="xl",
                                                        ),
                          
                          ])

group3 =  dbc.ButtonGroup([condition_load, renew, save_ideal_btn, page, page_2], id = 'buttongrp_3')

#CALLBACKS

###CONTROL CELLS FOR UNPUT#####  запрещают ввод в определенные ячейки       
@app.callback(
    Output("table-01", "columns"),
    Output("table-01", "dropdown_conditional"),
    [Input("table-01", "active_cell"),
    Input("table-01", "data"),
    ],
)
def tb01_status(cell, data): 
    ###new dropdowns lists###
    new_tbl =  functions2.sql_data_LIMIT500(data = 'dash_analogue', order_id = 'mpmin', limit = 1500) #sql_data(data = 'dash_analogue')
    new_tbl = new_tbl[['kalibrelm', 'markastali']]
    #print(new_tbl)
    new_tbl2 = functions2.sql_data_LIMIT500(data = 'dash_pipes', limit = 1500) #sql_data(data = 'dash_pipes')
    new_tbl2 = new_tbl2[['kalibrelm', 'markastali']]
    #print(new_tbl2)
    new_tbl3 = functions2.sql_data_LIMIT500(data = 'data_report', order_id = 'datetime', limit = 1500) #sql_data()
    #print(new_tbl3)
    new_tbl3.rename(columns = {"vtz_report1_1": "kalibrelm", 'vtz_report1_3': 'markastali'}, 
              inplace = True)
    new_tbl3 = new_tbl3[['kalibrelm', 'markastali']]
    #print(new_tbl3)
    new_tbl4 = functions2.sql_data_LIMIT500(data = 'integrazaya', order_id = 'timestamp', limit = 1500) #sql_data(data = 'integrazaya')
    new_tbl4 = new_tbl2[['kalibrelm', 'markastali']]
    #print(new_tbl4)
    new_tbl5 = pd.DataFrame()
    new_tbl5['kalibrelm'] = list_kalibr['kalibr']
    new_tbl5['markastali'] = list_marki['marki']
    #print(new_tbl5)
    
    Table_list = pd.concat([new_tbl, new_tbl2, new_tbl3, new_tbl4, new_tbl5])
    Table_list['markastali'] = Table_list['markastali'].astype('string')
    Table_list['kalibrelm'] = Table_list['kalibrelm'].astype('float')
    #Table_list = Table_list.dropna()

    list_marki_2 = Table_list['markastali'].unique()
    list_kalibr_2 = Table_list['kalibrelm'].unique()

    dropdown_conditional=[{
            'if': {
                'column_id': 'Значение',
                #'row_index': 2,
                'filter_query': '{Параметр} eq "Марка стали"'
            },
            'options': [
                            {'label': i, 'value': i}
                            for i in list_marki_2
                        ]
        },
       {
            'if': {
                'column_id': 'Значение',
                'filter_query': '{Параметр} eq "Калибр, мм"'
            },
            'options': [
                            {'label': i, 'value': i}
                            for i in list_kalibr_2
                        ]
        }]
    
    if len(data)>8:
        list_a = [4, 5]
    else:
        list_a = [4]
        
    #changed_id = [p['prop_id'] for p in dash.callback_context.triggered][0]

    if cell and cell['row'] !=None:
        
        if cell['row'] in list_a: #код оправки - не рададктируется
            columns= [
                {'name': 'Параметр', 'id': 'Параметр', "editable": False},
                {'name': 'Значение', 'id': 'Значение', "editable": False},]
            
        elif cell['row']==2 or cell['row']==3: #списки - марки стали и диаметр
            columns= [
                {'name': 'Параметр', 'id': 'Параметр', "editable": False},
                {'name': 'Значение', 'id': 'Значение', "editable": True, 'presentation': 'dropdown'},]
        else:
            columns= [
                {'name': 'Параметр', 'id': 'Параметр', "editable": False},
                {'name': 'Значение', 'id': 'Значение', "editable": True, 'type': 'numeric'},]      
    else:
        columns= [
                {'name': 'Параметр', 'id': 'Параметр', "editable": False},
                {'name': 'Значение', 'id': 'Значение', "editable": False},]
    return columns, dropdown_conditional        


@app.callback(
    Output("table-02", "columns"),
    [Input("table-02", "active_cell"),],
)
def tb02_status(cell):
    
    #changed_id = [p['prop_id'] for p in dash.callback_context.triggered][0]
    if cell and cell['row'] != None:
        if cell['row'] in [5,6,7,8,9,10]:
            columns= [
                {'name': 'Параметр', 'id': 'Параметр', "editable": False},
                {'name': 'Значение', 'id': 'Значение', "editable": False},]
        else:
            columns= [
                {'name': 'Параметр', 'id': 'Параметр', "editable": False},
                {'name': 'Значение', 'id': 'Значение', "editable": True, 'type': 'numeric'},]
    else:
        columns= [
                {'name': 'Параметр', 'id': 'Параметр', "editable": False},
                {'name': 'Значение', 'id': 'Значение', "editable": False},]         
    return columns      


@app.callback(
    Output("table-03", "columns"),
    [Input("table-03", "active_cell"),
    Input("table-03", "data"),
    ],
)
def tb03_status(cell, data):
    if len(data)>16:
        list_a = [0, 2,3,4, 6,7, 8, 10, 11, 12, 14, 15, 16, 18, 19, 20, 22, 23,  24, 26, 27]
    else:
        list_a = [0, 2, 4, 6, 8, 10, 12]
        
    #changed_id = [p['prop_id'] for p in dash.callback_context.triggered][0]
    if cell and cell['row'] !=None:

        if cell['row'] in list_a:
            columns= [
                {'name': 'Параметр', 'id': 'Параметр', "editable": False},
                {'name': 'Значение', 'id': 'Значение', "editable": False},]
            
        else:
            columns= [
                {'name': 'Параметр', 'id': 'Параметр', "editable": False},
                {'name': 'Значение', 'id': 'Значение', "editable": True, 'type': 'numeric'},]
        return columns  
    
    else:
        columns= [
                {'name': 'Параметр', 'id': 'Параметр', "editable": False},
                {'name': 'Значение', 'id': 'Значение', "editable": False},]
        return columns   
    

@app.callback(
    Output("table-04", "columns"),
    [Input("table-04", "active_cell"),
    Input("table-04", "data")
    ],
)
def tb04_status(cell, data):
    if len(data)>20:
        list_a = [0, 2,3,4, 6,7, 8, 10, 11, 12, 14, 15, 16, 18, 19, 20, 22, 23,  24, 
                   26, 27, 28, 30, 31, 32, 34, 35]
    else:
        list_a = [0, 2, 4, 6, 8, 10, 12, 14, 16]
    changed_id = [p['prop_id'] for p in dash.callback_context.triggered][0]

    if 'table-04.active_cell' in changed_id:

        if cell['row'] in list_a:
            columns= [
                {'name': 'Параметр', 'id': 'Параметр', "editable": False},
                {'name': 'Значение', 'id': 'Значение', "editable": False},]
        else:
            columns= [
                {'name': 'Параметр', 'id': 'Параметр', "editable": False},
                {'name': 'Значение', 'id': 'Значение', "editable": True, 'type': 'numeric'},]
        return columns  
    
    else:
        columns= [
                {'name': 'Параметр', 'id': 'Параметр', "editable": False},
                {'name': 'Значение', 'id': 'Значение', "editable": False},]   
        return columns

    
@app.callback(
    Output("loading_state", "fullscreen"),
    Output("loading_state", "style"),
    Input("model", "n_clicks"), 
    Input("ok", "n_clicks"),  
    Input("choose_analog_btn", "n_clicks"),
)
def loading_state(model, ok, analog):
    changed_id = [p['prop_id'] for p in dash.callback_context.triggered][0]
    if 'model' or "ok" or "analog" in changed_id:
        return True, {'display': 'block'}
    else:
        return False, {'display': 'none'}


@app.callback(
    [Output("modal", "is_open"), #всплывающее окно 1, появляется, если не хватает данных для модели
    Output("table-02", "data"), #2 - данные для таблицы 2 - отображается на экране
    Output("table-03", "data"), #3 - данные для таблицы 3 - отображается на экране
    Output("table-03", "style_data_conditional"), #4 - цвет строк в таблице по условию
    Output('hide', 'style'), #видимость графика со скоростями
    Output("modal2", "is_open"), #6 всплывающее окно 2, работа с Postgress - загрузка и удаление строк
    Output("table-04", "data"), #7 - данные для таблицы 4 - отображается на экране
    Output("table-04", "style_data_conditional"), #8 - цвет строк в таблице по условию
    Output("graph_1", "figure"), #9
    Output("table_from_sql", "children"),#10
    Output("select_table", "data"), #таблица с аналогами (появляется по условию)
    Output("table-01", "data"),#12
    Output("table-01", "style_data_conditional"), #13 - цвет строк в таблице по условию
    Output('store', 'data'), #текущий айди базы - (загруженный или посчитанный). В других случаях передается 0
    Output("loading-hidden-div", "children"),
    ],
    Input("table-01", "data"), #данные из вводных таблиц
    Input("table-02", "data"), #данные из вводных таблиц
    Input("table-03", "data"), #данные из вводных таблиц
    Input("model", "n_clicks"), #всплывающее окно 1 с предупреждением, если расчетн данные введены не полностью
    Input("close", "n_clicks"), #закрыть всплывающее окно 1
    Input("load", "n_clicks"), #всплывающее окно, позволяющее работа с БД
    Input("ok", "n_clicks"),  #подтвержает выбор строки из базы и перенос данных на экран
    Input("return", "n_clicks"), #закрывает всплывающее окно без выбора или удаления
    Input("del", "n_clicks"), #удаление строки
    Input("select_table", "selected_rows"), #выбор строки в таблице с аналогами. Отображает новый график, таблицу с зазорами и оборотами
    Input("clear", "n_clicks"),
    Input("choose_analog_btn", "n_clicks"), #кнопка для переноса анлога в расчет. Появляется при выборе строки в таблице аналога
    Input('store2', 'data'), #данные из другой функции. Нужны для передачи новых зазоров, 
    #оборотов и т.д. при выборе и сохранении аналога как расчета
    Input("table-04", "data"), #данные из вводных таблиц
    Input('store', 'data'),
    [State("modal", "is_open"),
    State("modal2", "is_open"),
    State("select_table", "data")]
)
def calc(data1, data2, data3, n_clicks, close_clicks, load_clicks, ok_clicks, 
         close2_clicks, del_clicks, row_sel_tb, clear_clicks, choose_analog_click, dt_analog, data4, previous_id,
         is_open, is_open2, select_table_data):    

    changed_id = [p['prop_id'] for p in dash.callback_context.triggered][0]
    if 'model' in changed_id:

        check_val_tb1 = sum(x['Значение'] is not None for x in data1)
        check_val_tb2 = sum(x['Значение'] is not None for x in data2)
        check_val_tb3 = sum(x['Значение'] is not None for x in data3)
        check_val_tb4 = sum(x['Значение'] is not None for x in data4)

        ##первоначальная - 
        if len(data2) == 11 and check_val_tb1 == 8 and check_val_tb2 == 11 and check_val_tb3 == 22 and check_val_tb4 == 27:

            model_df = pd.concat([pd.DataFrame(data1), pd.DataFrame(data2), pd.DataFrame(data3), pd.DataFrame(data4)])
            model_df = model_df.reset_index(drop = True)
            model_df = model_df.reset_index()
            model_data = df_for_model.merge(model_df, left_on='index', right_on='index', how = 'left')
            model_data = model_data[['name_properties', 'Значение']]
            model_data = model_data.rename(columns = {'Значение': 'new'}, inplace = False)
            model_data = model_data.fillna(0)
            #print(previous_id)
            
            #check data by condition 
            condition_1 = "float(ff(model_data[model_data['name_properties'] =='Калибр']['new'])) <= float(ff(model_data[model_data['name_properties'] =='Оправка МПМ']['new']))" 
            condition_2 = "float(ff(model_data[model_data['name_properties'] =='Толщина трубы']['new'])) > 30"
            condition_3 = "float(ff(model_data[model_data['name_properties'] =='Диаметр Трубы']['new'])) <100"
            condition_4 = "float(ff(model_data[model_data['name_properties'] =='Диаметр Трубы']['new'])) >1000"

            #условия, согласно которым введенные данные считаются неправильными и нуждаются в корректировке
            if eval(condition_1) == True or eval(condition_2) == True or eval(condition_3) == True or eval(condition_4) == True:
                #print('except by data condition')
                list_vals= [not is_open, no_update,  no_update,  no_update,  no_update, no_update, no_update,
                            no_update, no_update, no_update, no_update, no_update, no_update, no_update, no_update]
                return list_vals
              
            model_data['new'] = model_data['new'].astype('object')    
 
            try:
                
                
                model_ = model_vtz.model_v0(data_properties_VTZ=model_data)
                model_ = model_[['pred', 'name']]

                model_sql = head.merge(model_, left_on='name', right_on='name', how = 'left')
                model_sql = model_sql[['name', 'pred']].T
                model_sql.columns = model_sql.iloc[0]
                model_sql = model_sql.reset_index(drop = True)
                model_sql = model_sql.drop([0])

                sss_heattreatment=''
                for i in range(0, model_sql.shape[1]):
                    sss_heattreatment=sss_heattreatment + "%s,"
                sss_heattreatment=sss_heattreatment[:-1]

                for i in model_sql.columns:
                    if i != 'datetime':
                        try:
                            model_sql[i]=model_sql[i].astype(str)
                        except: 
                            model_sql[i]=model_sql[i].astype('object')
                        
                #по согласованию, всегда перезаписываем карту, если какие-либо расчеты были сделаны и расчет запустился заново        
                functions2.delete_row_by_ID(del_id=previous_id)##удаляем предыдущую версию, оставляем только перезаписанную
                #пересохраняем со старым ID
                model_sql['datetime']=model_sql['datetime'].astype('datetime64[ns]') + timedelta(hours=3)
                model_sql['scenarioid'] = previous_id

                down_status = functions2.save_to_sql(sql_base = model_sql, sss_heattreatment = sss_heattreatment)
                #print(down_status)
                current_id = ff(model_sql['scenarioid'])
                #print('resave', current_id)

                df_select = functions2.data_by_ID(track_id = current_id)
                df_select = df_select.round(1)
                try:
                    df_select['vtz_report1_6'] = round((df_select['vtz_report1_6'].astype('float') * df_select['cuts'].astype('float')), 0)
                    #print('ok')
                except:
                    df_select['vtz_report1_6'] = df_select['vtz_report1_6']

                y_data_new = df_select[['vtz_report2_155', 'vtz_report2_156', 'vtz_report2_157', 'vtz_report2_158', 
                          'vtz_report2_159', 'vtz_report2_160', 'vtz_report2_161']].values.tolist()
                y_data_float = list(np.float_(y_data_new))
                df_select2 = df_select[['vtz_report2_1', 'vtz_report2_3', 'vtz_report1_1', 'vtz_report1_2']]
                check_list = df_select2.values.tolist()

                list_vals= [is_open, display_tb(df, df2 = df_select)[1].to_dict('records'), 
                            display_tb(df, df_select)[2].to_dict('records'), style_table3,
                            {'display': 'block'}, no_update, 
                            display_tb(df, df2 = df_select)[3].to_dict('records'), style_table4, graph_1(x_data, y_data_float[0]), #no_update, 
                            create_analog_tabl(slice_tb(check_list, current_id)),
                            no_update, display_tb(df, df2 = df_select)[0].to_dict('records'),
                           style_table1_2, current_id, []]
            except:
                #print('except')
                list_vals= [not is_open, no_update,  no_update,  no_update,  no_update, no_update, no_update,
                            no_update, no_update, no_update, no_update, no_update, no_update, no_update, no_update]
                
    
            return list_vals

        elif len(data2) == 5 and check_val_tb1 == 7 and check_val_tb2 == 5 and check_val_tb3 == 8 and check_val_tb4 == 9:

            model_df = pd.concat([pd.DataFrame(data1), pd.DataFrame(data2), pd.DataFrame(data3), pd.DataFrame(data4)])
            model_df = model_df.reset_index(drop = True)
            model_df = model_df.reset_index()

            model_data = df_for_model.merge(model_df, left_on='index_new', right_on='index', how = 'left')
            model_data = model_data[['name_properties', 'Значение']]
            model_data = model_data.rename(columns = {'Значение': 'new'}, inplace = False)
            model_data = model_data.fillna(0)
            model_data.to_excel("data_input_model.xlsx")
            model_data['new'] = model_data['new'].astype('object')
            
            try:
                model_ = model_vtz.model_v0(data_properties_VTZ=model_data)
                model_ = model_[['pred', 'name']]

                model_sql = head.merge(model_, left_on='name', right_on='name', how = 'left')
                model_sql = model_sql[['name', 'pred']].T
                model_sql.columns = model_sql.iloc[0]
                model_sql = model_sql.reset_index(drop = True)
                model_sql = model_sql.drop([0])
                #print(model_sql)

                sss_heattreatment=''
                for i in range(0, model_sql.shape[1]):
                    sss_heattreatment=sss_heattreatment + "%s,"
                sss_heattreatment=sss_heattreatment[:-1]

                for i in model_sql.columns:
                    if i != 'datetime':
                        try:
                            model_sql[i]=model_sql[i].astype(str)
                        except: 
                            model_sql[i]=model_sql[i].astype('object')

                model_sql['datetime']=model_sql['datetime'].astype('datetime64') + timedelta(hours=3)
            
                down_status = functions2.save_to_sql(sql_base = model_sql, sss_heattreatment = sss_heattreatment)
                #print(down_status)
                current_id = ff(model_sql['scenarioid'])
                #print(current_id)

                df_select = functions2.data_by_ID(track_id = current_id)
                df_select = df_select.round(1)
                try:
                    df_select['vtz_report1_6'] = round((df_select['vtz_report1_6'].astype('float') * df_select['cuts'].astype('float')), 0)
                except:
                    df_select['vtz_report1_6'] = df_select['vtz_report1_6']

                y_data_new = df_select[['vtz_report2_155', 'vtz_report2_156', 'vtz_report2_157', 'vtz_report2_158', 
                          'vtz_report2_159', 'vtz_report2_160', 'vtz_report2_161']].values.tolist()
                y_data_float = list(np.float_(y_data_new))
                df_select2 = df_select[['vtz_report2_1', 'vtz_report2_3', 'vtz_report1_1', 'vtz_report1_2']]
                check_list = df_select2.values.tolist()

                list_vals= [is_open, display_tb(df, df2 = df_select)[1].to_dict('records'), 
                            display_tb(df, df_select)[2].to_dict('records'), style_table3,
                            {'display': 'block'}, no_update, 
                            display_tb(df, df2 = df_select)[3].to_dict('records'), style_table4, graph_1(x_data, y_data_float[0]), #no_update, 
                            create_analog_tabl(slice_tb(check_list, current_id)),
                            no_update, display_tb(df, df2 = df_select)[0].to_dict('records'),
                           style_table1_2, current_id, []]
            except:
                #print('except')
                list_vals= [not is_open, no_update,  no_update,  no_update,  no_update, no_update, no_update,
                            no_update, no_update, no_update, no_update, no_update, no_update, no_update, no_update]
                
            
            return list_vals  
        
        else:
            return not is_open, no_update,  no_update,  no_update,  no_update, no_update, no_update,\
        no_update, no_update, no_update, no_update, no_update, no_update, no_update, []
                       
    elif 'close' in changed_id:
        return not is_open, no_update,  no_update,  no_update,  no_update, no_update, no_update,\
    no_update, no_update, no_update, no_update, no_update, no_update, no_update, []
    
    elif 'load' in changed_id:
        #при открытии вплывающего окна заново подгружаем данные из БД
        pg1 = choose_table_create(functions2.sql_data_LIMIT500(data = 'data_report', order_id = 'datetime', limit = 300))

        return is_open, no_update,  no_update,  no_update,  no_update, not is_open2,\
    no_update, no_update, no_update, no_update, pg1.to_dict('record'), no_update, no_update, no_update,\
    []
    
    elif 'return' in changed_id:
        return no_update, no_update,  no_update,  no_update,  no_update, not is_open2,\
    no_update, no_update, no_update, no_update, no_update, no_update, no_update, no_update, []
    
    elif 'ok' in changed_id:
        
        pg = choose_table_create(functions2.sql_data_LIMIT500(data = 'data_report', order_id = 'datetime', limit = 300))
        current_id = ff(pd.DataFrame(select_table_data)['scenarioid'][row_sel_tb]) #текущий айди,  который будет передаваться между функциями 
        #print(row_sel_tb)
        #print(current_id)
        df_select = functions2.data_by_ID(track_id = ff(pg['scenarioid'][row_sel_tb]))

        y_data_new = df_select[['vtz_report2_155', 'vtz_report2_156', 'vtz_report2_157', 'vtz_report2_158', 
                  'vtz_report2_159', 'vtz_report2_160', 'vtz_report2_161']].values.tolist()

        y_data_float = list(np.float_(y_data_new))
        
          
        df_select2 = df_select[['vtz_report2_1', 'vtz_report2_3', 'vtz_report1_1', 'vtz_report1_2']]
        check_list = df_select2.values.tolist()
        #print(check_list)

        try:
            df_select['vtz_report1_6'] = round((df_select['vtz_report1_6'].astype('float') * df_select['cuts'].astype('float')), 0)
        except:
            df_select['vtz_report1_6'] = df_select['vtz_report1_6']
        
        list_vals= [is_open, display_tb(df, df2 = df_select)[1].to_dict('records'), 
                    display_tb(df, df2 = df_select)[2].to_dict('records'), style_table3,
                    {'display': 'block'}, not is_open2, 
                    display_tb(df, df2 = df_select)[3].to_dict('records'), style_table4, 
                    graph_1(x_data, y_data_float[0]),
                    create_analog_tabl(slice_tb(check_list, current_id)),
                    no_update, display_tb(df, df2 = df_select)[0].to_dict('records'),
                   style_table1_2,
                   current_id, []]
        

        return list_vals
    
    elif 'del' in changed_id:
        #удаляем строку по айди
        if row_sel_tb != None:
            pg = choose_table_create(functions2.sql_data_LIMIT500(data = 'data_report', order_id = 'datetime', limit = 600))
            #print(pg['scenarioid'][row_sel_tb])
            #print(ff(pg['scenarioid'][row_sel_tb]))
            functions2.delete_row_by_ID(del_id=ff(pg['scenarioid'][row_sel_tb]))
        
        return is_open, no_update,  no_update,  no_update,  no_update, not is_open2,\
    no_update, no_update, no_update, no_update, no_update, no_update, no_update, no_update, []
    
    elif 'clear' in changed_id:

        current_id = '0'
        style3 = (
        [{'if': {'row_index': d},
            'backgroundColor': "#F5F5F5",
            'font-weight': 'bold',
        } for d in [0, 2, 4, 6, 8, 10, 12]])
        style4 =(
        [{'if': {'row_index': d},
            'backgroundColor': "#F5F5F5",
            'font-weight': 'bold',
        } for d in [0, 2, 4, 6, 8, 10, 12, 14, 16]
        ])
     
        return is_open, display_tb(df, df2 = pd.DataFrame())[1].iloc[0:5].to_dict('records'),\
    display_tb(df, df2 = pd.DataFrame())[2].iloc[[0,1,4,5,8,9,12,13,16,17,20,21,24,25,28]].to_dict('records'),\
    style3,  {'display': 'none'}, no_update,\
    display_tb(df, df2 = pd.DataFrame())[3].iloc[[0,1,4,5,8,9,12,13,16,17,20,21,24,25,28,29,32,33]].to_dict('record'),\
    style4, no_update, dash_table.DataTable(id='table_analog'), no_update,\
    display_tb(df, df2 = pd.DataFrame())[0].to_dict('records'), style_table1_1, current_id, []

    elif 'choose_analog_btn' in changed_id:
        #пересохраняем аналог, как новые данные (меняем в старых данных зазоры, обороты, скорости) 
        df_select_analog = pd.DataFrame.from_dict(dt_analog[0])
        #print(df_select_analog['scenarioid'])

        df_select_analog.loc[0, ['vtz_report2_43', 'vtz_report2_44', 'vtz_report2_45', 'vtz_report2_46',
                                        'vtz_report2_47', 'vtz_report2_48', 'vtz_report2_49',
                                         'vtz_report2_106', 'vtz_report2_107', 'vtz_report2_108', 'vtz_report2_109',
                                         'vtz_report2_110', 'vtz_report2_111', 'vtz_report2_112', 'vtz_report2_113',
                            'vtz_report2_114']]  = dt_analog[1]

        df_select_analog.loc[0, ['vtz_report2_36', 'vtz_report2_37', 'vtz_report2_38', 'vtz_report2_39',
                                        'vtz_report2_40', 'vtz_report2_41', 'vtz_report2_42',
                                         'vtz_report2_95', 'vtz_report2_96', 'vtz_report2_97', 'vtz_report2_98',
                                         'vtz_report2_99', 'vtz_report2_100', 'vtz_report2_101', 'vtz_report2_102',
                            'vtz_report2_103']] = dt_analog[2]
        df_select_analog.loc[0, ['vtz_report2_155', 'vtz_report2_156', 'vtz_report2_157', 'vtz_report2_158', 
                  'vtz_report2_159', 'vtz_report2_160', 'vtz_report2_161']] = dt_analog[4]
        
        df_select_analog.loc[0, ['vtz_report2_164', 'vtz_report2_165', 'vtz_report2_166', 'vtz_report2_167', 
                  'vtz_report2_168', 'vtz_report2_169', 'vtz_report2_170']] = dt_analog[5]
        
        #пересчитываем новое время и id
        now = datetime.now() + timedelta(hours=3)
        df_select_analog["datetime"] = now.strftime("%Y-%m-%d %H:%M:%S")

#         df_select_analog['scenarioid'] = str(ff(df_select_analog["vtz_report2_1"])) + "_" + \
#                                         str(ff(df_select_analog["vtz_report2_3"])) + "_" + \
#                                         str(ff(df_select_analog["vtz_report1_12"])) + "_" + \
#                                         str(ff(df_select_analog["vtz_report1_1"])) + "_" +\
#                                         str(ff(df_select_analog["vtz_report1_3"])) + "_" +\
#                                         str(ff(df_select_analog["datetime"]))

        #диаметр, стенка, оправка, калибр, марка, дата
        df_select_analog['source'][0] = str(dt_analog[3]) #добавляем источник
        
        df_select_analog['vtz_report2_130']=round(701.6601671630542-0.4069175131244827*\
                                                  float(df_select_analog['vtz_report2_129'])+\
            0.0441577157201906*float(df_select_analog['vtz_report2_155'])+\
            0.04415771572019076*float(df_select_analog['vtz_report2_156'])+\
            0.008953055135952077*float(df_select_analog['vtz_report2_157'])+\
            0.0015132335923520382*float(df_select_analog['vtz_report2_158'])+\
            0.004739170513118333*float(df_select_analog['vtz_report2_159'])+\
            0.0672449789942485*float(df_select_analog['vtz_report2_160'])+\
            0.013632662532926467*float(df_select_analog['vtz_report2_161']),0)
        
        sss_heattreatment=''

        for i in range(0, df_select_analog.shape[1]):
            sss_heattreatment=sss_heattreatment + "%s,"
        sss_heattreatment=sss_heattreatment[:-1]
        
        for i in df_select_analog.columns:
            if i != 'datetime':
                try:
                    df_select_analog[i]=df_select_analog[i].astype(str)
                except: 
                    df_select_analog[i]=df_select_analog[i].astype('object')

        df_select_analog['datetime']=df_select_analog['datetime'].astype('datetime64')              
        previous_id = previous_id ##предыдущий айдишник - рассчитанный или загруженный
        #print(previous_id)
        
        df_select_analog2 = df_select_analog[['vtz_report2_1', 'vtz_report2_3', 'vtz_report1_1', 'vtz_report1_2']]
        check_list = df_select_analog2.values.tolist()
        current_id = ff(df_select_analog['scenarioid'])
        
        #print(previous_id, current_id)
        y_data_new = df_select_analog[['vtz_report2_155', 'vtz_report2_156', 'vtz_report2_157', 'vtz_report2_158', 
                  'vtz_report2_159', 'vtz_report2_160', 'vtz_report2_161']].values.tolist()

        y_data_float = list(np.float_(y_data_new))
        
        
        functions2.delete_row_by_ID(del_id=previous_id)##удаляем предыдущую версию, оставляем только перезаписанную
        down_status = functions2.save_to_sql(sql_base = df_select_analog, sss_heattreatment = sss_heattreatment)
        #print(down_status)
        
        list_vals= [is_open, display_tb(df, df2 = df_select_analog)[1].to_dict('records'), 
                    display_tb(df, df2 = df_select_analog)[2].to_dict('records'), style_table3,
                    {'display': 'block'}, is_open2, 
                    display_tb(df, df2 = df_select_analog)[3].to_dict('records'), style_table4, 
                    graph_1(x_data, y_data_float[0]), 
                    create_analog_tabl(slice_tb(check_list, current_id)),
                    no_update, display_tb(df, df2 = df_select_analog)[0].to_dict('records'),
                   style_table1_2,
                   current_id, []]
        return list_vals 
    
    else:
        return is_open, no_update,  no_update,  no_update,  no_update, no_update,\
    no_update, no_update, no_update, no_update, no_update, no_update, no_update, no_update, []


    
@app.callback(
    Output("div_analog_table", "children"), #таблица с оборотами и зазорами
    #Output("clear", "n_clicks"),
    Output('hide', 'children'), #обновление графика (добавляем аналоги)
    Output('choose_analog_btn', 'style'), #видимость кнопки с возможностью переноса анлога
    Output('store2', 'data'),#данные для передачи в другую функцию (выбранныя строка, обороты, зазаоры и т.д.)
    Output('save_analogpipes_div', 'style'),
    [
    #Input("table_analog", "derived_virtual_selected_rows"), #выбор строки
    Input("table_analog", "selected_rows"), #выбор строки
    Input("table_analog", "data"),
    Input('store', 'data')], #айди ткущей строки из БД
    #Input("clear", "n_clicks"),
    
    )
def update_new_tb(row_out, table_data, data):
    #функция, возвращающая таблицу с оборотами и зазорами аналога + обновляющая график (красная линия - аналог)

    if row_out:
        #print(row_out)
        if row_out==[]:
            return dash_table.DataTable(id='zazor_table'), no_update, {'display': 'none'}, None, {'display': 'none'}

        elif row_out!=[]:
            df_select = functions2.data_by_ID(track_id = data)
            y_data_new = df_select[['vtz_report2_155', 'vtz_report2_156', 'vtz_report2_157', 'vtz_report2_158', 
                  'vtz_report2_159', 'vtz_report2_160',  'vtz_report2_161']].values.tolist()
            y_data_float = list(np.float_(y_data_new))
            
            Kalibr = ff(functions2.data_by_ID(data)['vtz_report2_6'])
            
            
            diam_fact_mpm_new = functions2.data_by_ID(data)[['vtz_report2_29', 'vtz_report2_30', 'vtz_report2_31', 
                                                     'vtz_report2_32', 'vtz_report2_33', 
                                                     'vtz_report2_34', 'vtz_report2_35', 'vtz_report2_85', 
                                                     'vtz_report2_86', 'vtz_report2_87', 'vtz_report2_88', 
                                                     'vtz_report2_89', 'vtz_report2_90', 'vtz_report2_91', 
                                                     'vtz_report2_92', 'vtz_report2_93']]
            
            #по нужной строке из созданной таблице берем айди (не отображ на экране)
                                   
            if table_data[int(row_out[0])].get('Источник данных') == 'Данные со стана' or table_data[int(row_out[0])].get('Источник данных') == 'Данные со стана (интеграция)':
                #берем уникальную строку по дате + айди
                if table_data[int(row_out[0])].get('Источник данных') == 'Данные со стана':
                    dt = functions2.data_by_ID(table_data[int(row_out[0])].get('Data'), 
                                  data = 'dash_analogue', id_ = "mpmin")
                    dt = dt[dt['idpipesinuque'] == table_data[int(row_out[0])].get('ID')]
                
                else:
                    dt = functions2.data_by_ID(table_data[int(row_out[0])].get('ID'), data = 'integrazaya', id_ = "id")
                    
                #диаметры, зазоры и обороты аналога для отображения и пересчета                    
                diam_fact_mpm = dt[['diametrvalkovmmklet1mpm',
                                       'diametrvalkovmmklet2mpm', 'diametrvalkovmmklet3mpm',
                                       'diametrvalkovmmklet4mpm', 'diametrvalkovmmklet5mpm',
                                       'diametrvalkovmmklet6mpm', 'diametrvalkovmmklet7mpm',
                                       'diametrvalkovmmklet1esm', 'diametrvalkovmmklet2esm',
                                       'diametrvalkovmmklet3esm', 'diametrvalkovmmklet4esm',
                                       'diametrvalkovmmklet5esm', 'diametrvalkovmmklet6esm',
                                       'diametrvalkovmmklet7esm', 'diametrvalkovmmklet8esm',
                                       'diametrvalkovmmklet9esm']]
                
                zazor = dt[['zazorvalkovmmklet1mpm', 'zazorvalkovmmklet2mpm',
                           'zazorvalkovmmklet3mpm', 'zazorvalkovmmklet4mpm',
                           'zazorvalkovmmklet5mpm', 'zazorvalkovmmklet6mpm',
                           'zazorvalkovmmklet7mpm', 'zazorvalkovmmklet1esm',
                           'zazorvalkovmmklet2esm', 'zazorvalkovmmklet3esm',
                           'zazorvalkovmmklet4esm', 'zazorvalkovmmklet5esm',
                           'zazorvalkovmmklet6esm', 'zazorvalkovmmklet7esm',
                           'zazorvalkovmmklet8esm', 'zazorvalkovmmklet9esm']]
                oborot = dt[['skorostvalkovklet1obminmpm',
                               'skorostvalkovklet2obminmpm', 'skorostvalkovklet3obminmpm',
                               'skorostvalkovklet4obminmpm', 'skorostvalkovklet5obminmpm',
                               'skorostvalkovklet6obminmpm', 'skorostvalkovklet7obminmpm',
                                'skorostkletiobminklet1esm', 'skorostkletiobminklet2esm',
                               'skorostkletiobminklet3esm', 'skorostkletiobminklet4esm',
                               'skorostkletiobminklet5esm', 'skorostkletiobminklet6esm',
                               'skorostkletiobminklet7esm', 'skorostkletiobminklet8esm',
                               'skorostkletiobminklet9esm']]


            elif table_data[int(row_out[0])].get('Источник данных') == 'Фактические настройки': #по источнику данных определяем нужну таблицу в Postgress
                #берем уникальную строку по дате + айди
                dt = functions2.data_by_ID(table_data[int(row_out[0])].get('ID'), 
                              data = 'dash_pipes', id_ = "id")
                
                diam_fact_mpm = dt[['diametrvalkovmmklet1mpm',
                                   'diametrvalkovmmklet2mpm', 'diametrvalkovmmklet3mpm',
                                   'diametrvalkovmmklet4mpm', 'diametrvalkovmmklet5mpm',
                                   'diametrvalkovmmklet6mpm', 'diametrvalkovmmklet7mpm', 
                                    'diametrvalkovmmklet1esm', 'diametrvalkovmmklet2esm',
                                   'diametrvalkovmmklet3esm', 'diametrvalkovmmklet4esm',
                                   'diametrvalkovmmklet5esm', 'diametrvalkovmmklet6esm',
                                   'diametrvalkovmmklet7esm', 'diametrvalkovmmklet8esm',
                                   'diametrvalkovmmklet9esm',]]
                
                zazor = dt[['zazorvalkovmmklet1mpm',
                           'zazorvalkovmmklet2mpm', 'zazorvalkovmmklet3mpm',
                           'zazorvalkovmmklet4mpm', 'zazorvalkovmmklet5mpm',
                           'zazorvalkovmmklet6mpm', 'zazorvalkovmmklet7mpm',
                            'zazorvalkovmmklet1esm', 'zazorvalkovmmklet2esm',
                           'zazorvalkovmmklet3esm', 'zazorvalkovmmklet4esm',
                           'zazorvalkovmmklet5esm', 'zazorvalkovmmklet6esm',
                           'zazorvalkovmmklet7esm', 'zazorvalkovmmklet8esm',
                           'zazorvalkovmmklet9esm']]
                oborot = dt[['skorostvalkovklet1obminmpm', 'skorostvalkovklet2obminmpm',         
                           'skorostvalkovklet3obminmpm', 'skorostvalkovklet4obminmpm',
                           'skorostvalkovklet5obminmpm', 'skorostvalkovklet6obminmpm',
                           'skorostvalkovklet7obminmpm', 
                            'skorostkletiobminklet1esm',
                           'skorostkletiobminklet2esm', 'skorostkletiobminklet3esm',
                           'skorostkletiobminklet4esm', 'skorostkletiobminklet5esm',
                           'skorostkletiobminklet6esm', 'skorostkletiobminklet7esm',
                               'skorostkletiobminklet8esm', 'skorostkletiobminklet9esm']]
  
            else:
                #берем уникальную строку по дате + айди
                dt = functions2.data_by_ID(table_data[int(row_out[0])].get('ID'))
                #берем уникальную строку по дате + айди
                diam_fact_mpm = dt[['vtz_report2_29', 'vtz_report2_30', 'vtz_report2_31', 
                                                     'vtz_report2_32', 'vtz_report2_33', 
                                                     'vtz_report2_34', 'vtz_report2_35', 'vtz_report2_85', 
                                                     'vtz_report2_86', 'vtz_report2_87', 'vtz_report2_88', 
                                                     'vtz_report2_89', 'vtz_report2_90', 'vtz_report2_91', 
                                                     'vtz_report2_92', 'vtz_report2_93']]
                
            
                zazor = dt[['vtz_report2_43', 'vtz_report2_44', 'vtz_report2_45', 'vtz_report2_46',
                                        'vtz_report2_47', 'vtz_report2_48', 'vtz_report2_49',
                                         'vtz_report2_106', 'vtz_report2_107', 'vtz_report2_108', 'vtz_report2_109',
                                         'vtz_report2_110', 'vtz_report2_111', 'vtz_report2_112', 'vtz_report2_113',
                            'vtz_report2_114']]
                oborot = dt[['vtz_report2_36', 'vtz_report2_37', 'vtz_report2_38', 'vtz_report2_39',
                                        'vtz_report2_40', 'vtz_report2_41', 'vtz_report2_42',
                                         'vtz_report2_95', 'vtz_report2_96', 'vtz_report2_97', 'vtz_report2_98',
                                         'vtz_report2_99', 'vtz_report2_100', 'vtz_report2_101', 'vtz_report2_102',
                            'vtz_report2_103']]

            #округляем данные
            diam_fact_mpm = (pd.to_numeric(ff(diam_fact_mpm.values.tolist()), downcast="float", errors='coerce')).astype(float).round(1)
            diam_fact_mpm_new = (pd.to_numeric(ff(diam_fact_mpm_new.values.tolist()), downcast="float", errors='coerce')).astype(float).round(1)
            oborot = (pd.to_numeric(ff(oborot.values.tolist()), downcast="float", errors='coerce')).astype(float).round(0)
            zazor = (pd.to_numeric(ff(zazor.values.tolist()), downcast="float", errors='coerce')).astype(float).round(1)

            skor_zazor = return_speed(zazor= zazor[0:7], #аналог
                    oborot= oborot[0:7],
                    Kalibr=float(Kalibr), #аналог
                    diam_fact_mpm=diam_fact_mpm[0:7], #аналог
                    diam_fact_mpm_new= diam_fact_mpm_new[0:7], #пользователь
                    data_prop=data_prop)

            y_data =  y_data_float[0] #данные для графика - модель
            z_data = skor_zazor[0] #данные для графика - аналоги
            oborot_esm = np.round((diam_fact_mpm[7:]*oborot[7:])/diam_fact_mpm_new[7:], 0) #пересчет оборотов для esm

            tabl_zazor = tables_zazory(df, df_select, skor = (skor_zazor[1] + oborot_esm.tolist()), zazor = zazor)
            
            return zazor_table(tabl_zazor), dbc.Card(dcc.Graph(figure =  graph_2(x_data, y_data, z_data), 
                                              style = {'height': "22vw", 'align': 'center'}, 
                                                                              id = 'graph_1'
                                                                              
                                             ), body = True, color="light"), {'display': 'block'}, [df_select.to_dict('records'), 
                                                                                                    zazor, 
                                                                                                    (skor_zazor[1] + oborot_esm.tolist()),
                                                                                                   table_data[int(row_out[0])].get('Источник данных'),
                                                                                                   skor_zazor[0],
                                                                                                   skor_zazor[2]], {'display': 'block'}
        
        return dash_table.DataTable(id='zazor_table'), no_update, {'display': 'none'}, None, {'display': 'none'}
    else:
        return dash_table.DataTable(id='zazor_table'), no_update, {'display': 'none'}, None, {'display': 'none'}


    
@app.callback(
    Output("download", "data"),
    [
     Input("item-MPM", "n_clicks"), 
     Input("item-PS", "n_clicks"),
     Input('store', 'data')
    ])

def update_download_href(n_clicks, n_clicks2, current_id):
##функция, позволяющая скачивать отчеты в формате эксель. Передаем тип отчета (МПМ или ПС), 
#и id, по которому подтягиваем данные из Postgres
    changed_id = [p['prop_id'] for p in dash.callback_context.triggered][0]

    if current_id is None:
        current_id = '0'
        
    if 'item-MPM' in changed_id:
        df_report = functions2.data_by_ID(track_id = current_id)

        if df_report.empty == True: #если данных нет, то возвращаем 0
            series_obj = pd.Series([0]*df_report.shape[1], 
                        index=df_report.columns)
  
            df_report = df_report.append(series_obj,
                                    ignore_index=True)
    
        replace_ = df_report.to_dict('list')
        replace_ = {i:str(j[0]) for i,j in replace_.items()}

        template(replace_, card_name ="MPM.xlsx")
        return send_file('MPM.xlsx')    
    
    elif 'item-PS' in changed_id:

        df_report = functions2.data_by_ID(track_id = current_id)
        if df_report.empty == True: #если данных нет, то возвращаем 0
                series_obj = pd.Series([0]*df_report.shape[1], 
                            index=df_report.columns )

                df_report = df_report.append(series_obj,
                                        ignore_index=True)
        replace_ = df_report.to_dict('list')
        replace_ = {i:str(j[0]) for i,j in replace_.items() }

        template(replace_, card_name ="PS.xlsx")
        return send_file('PS.xlsx')

###calback resolve problems with filter on 2 and other pages###
@app.callback(
    Output('table_analog', "page_current"),
    Input('table_analog', "filter_query"))
def update_table_filter_1(filter):
    return 0

@app.callback(
    Output('select_table', "page_current"),
    Input('select_table', "filter_query"))
def update_table_filter_2(filter):
    return 0

###save as ideal pipes###
@app.callback(Output('load_resultpg1_2', 'children'), #display last update time
             Output('modal_analog_page1', 'is_open'),
             Output('modal_analog2_page1', 'is_open'),
             Output('load_resultpg1_1', 'children'),
             Input('save_ideal_btn_pg1', 'n_clicks'),
             Input('closeanalogmodal_page1', 'n_clicks'),
             Input('rejectsave_page1', 'n_clicks'),
             Input('oksave_page1', 'n_clicks'),
             State("table_analog", "selected_rows"),
             State("table_analog", "data"),
             State('modal_analog_page1', 'is_open'),
             State('modal_analog2_page1', 'is_open'),
             prevent_initial_call=True)
def page1saveideal(btn_save, close, reject, resave, row_out, table_data, is_open, is_open2):  #(btn_save, close, reject, save, row_id, is_open , is_open2):
#     print('save') 
#     print(btn_save)
#     print(row_out)
    old_colnames = ['vtz_report2_1', 'vtz_report2_3', 'vtz_report1_1', 'vtz_report1_2' ,'vtz_report1_3',
                    'vtz_report2_29', 'vtz_report2_30', 'vtz_report2_31', 'vtz_report2_32', 'vtz_report2_33', 
                    'vtz_report2_34', 'vtz_report2_35', 'vtz_report2_85', 'vtz_report2_86', 'vtz_report2_87', 'vtz_report2_88', 
                    'vtz_report2_89', 'vtz_report2_90', 'vtz_report2_91', 'vtz_report2_92', 'vtz_report2_93',
                    'vtz_report2_43', 'vtz_report2_44', 'vtz_report2_45', 'vtz_report2_46','vtz_report2_47', 'vtz_report2_48', 'vtz_report2_49',
                    'vtz_report2_106', 'vtz_report2_107', 'vtz_report2_108', 'vtz_report2_109',
                    'vtz_report2_110', 'vtz_report2_111', 'vtz_report2_112', 'vtz_report2_113', 'vtz_report2_114',
                    'vtz_report2_36', 'vtz_report2_37', 'vtz_report2_38', 'vtz_report2_39',
                    'vtz_report2_40', 'vtz_report2_41', 'vtz_report2_42',
                    'vtz_report2_95', 'vtz_report2_96', 'vtz_report2_97', 'vtz_report2_98',
                    'vtz_report2_99', 'vtz_report2_100', 'vtz_report2_101', 'vtz_report2_102', 'vtz_report2_103', 'vtz_report1_46'] 
                     
    new_colnames = ['diametrtruby', 'stenkatruby', 'kalibrelm', 'diametropravkimmmpm' ,'markastali',
                    'diametrvalkovmmklet1mpm','diametrvalkovmmklet2mpm', 'diametrvalkovmmklet3mpm',
                    'diametrvalkovmmklet4mpm', 'diametrvalkovmmklet5mpm','diametrvalkovmmklet6mpm', 'diametrvalkovmmklet7mpm', 
                    'diametrvalkovmmklet1esm', 'diametrvalkovmmklet2esm', 'diametrvalkovmmklet3esm', 'diametrvalkovmmklet4esm',
                    'diametrvalkovmmklet5esm', 'diametrvalkovmmklet6esm','diametrvalkovmmklet7esm', 'diametrvalkovmmklet8esm',
                    'diametrvalkovmmklet9esm','zazorvalkovmmklet1mpm','zazorvalkovmmklet2mpm', 'zazorvalkovmmklet3mpm',
                    'zazorvalkovmmklet4mpm', 'zazorvalkovmmklet5mpm','zazorvalkovmmklet6mpm', 'zazorvalkovmmklet7mpm',
                     'zazorvalkovmmklet1esm', 'zazorvalkovmmklet2esm','zazorvalkovmmklet3esm', 'zazorvalkovmmklet4esm',
                    'zazorvalkovmmklet5esm', 'zazorvalkovmmklet6esm', 'zazorvalkovmmklet7esm', 'zazorvalkovmmklet8esm',
                    'zazorvalkovmmklet9esm', 'skorostvalkovklet1obminmpm', 'skorostvalkovklet2obminmpm',         
                    'skorostvalkovklet3obminmpm', 'skorostvalkovklet4obminmpm','skorostvalkovklet5obminmpm', 'skorostvalkovklet6obminmpm',
                    'skorostvalkovklet7obminmpm', 'skorostkletiobminklet1esm','skorostkletiobminklet2esm', 'skorostkletiobminklet3esm',
                    'skorostkletiobminklet4esm', 'skorostkletiobminklet5esm','skorostkletiobminklet6esm', 'skorostkletiobminklet7esm',
                    'skorostkletiobminklet8esm', 'skorostkletiobminklet9esm', 'gilza']
    
    changed_id = [p['prop_id'] for p in dash.callback_context.triggered][0]
    # print(changed_id)
    if 'save_ideal_btn_pg1' in changed_id:
        if row_out and row_out != []:
            text = 'check'
            # print(table_data[int(row_out[0])].get('Источник данных'))
            if table_data[int(row_out[0])].get('Источник данных') == 'Данные со стана':
                dt = functions2.data_by_ID(table_data[int(row_out[0])].get('Data'), 
                                  data = 'dash_analogue', id_ = "mpmin")
                dt = dt[dt['idpipesinuque'] == table_data[int(row_out[0])].get('ID')]
            elif table_data[int(row_out[0])].get('Источник данных') == 'Данные со стана (интеграция)':
                #берем уникальную строку по дате + айди
                dt = functions2.data_by_ID(table_data[int(row_out[0])].get('ID'), data = 'integrazaya', id_ = "id")
                    
            elif table_data[int(row_out[0])].get('Источник данных') == 'Фактические настройки': #по источнику данных определяем нужну таблицу в Postgress
                #берем уникальную строку по дате + айди
                dt = functions2.data_by_ID(table_data[int(row_out[0])].get('ID'), 
                              data = 'dash_pipes', id_ = "id")              
            else:
                #берем уникальную строку по дате + айди
                dt = functions2.data_by_ID(table_data[int(row_out[0])].get('ID'))
#                 rename columns to save
                dt.rename(columns={i:j for i,j in zip(old_colnames,new_colnames)}, inplace=True)

            list_for_check = dt[['diametrtruby', 'stenkatruby', 'kalibrelm', 
                                                    'diametropravkimmmpm' ,'markastali',]]
            
            list_check = ff(list_for_check.values.tolist()) 
            list_1 = [float(i) for i in list_check[0:4]]
            list_1.append(list_check[4]) 
            
            ####проверка на наличие####
            sql_query2 = 'WHERE diametrtruby = (%s) AND stenkatruby = (%s) AND  kalibrelm = (%s) AND diametropravkimmmpm = (%s) AND markastali = (%s) ORDER BY ctid DESC LIMIT 500'
            dfquery = functions2.data_query(sql_query2, list_1,  data = 'dash_pipes')
            
            if len(dfquery) > 0:
                text ="Данные с диаметром: {}, стенкой: {}, калибром: {}, диаметром оправки: {} и маркой стали: {} уже существуют".format(list_1[0],
                                                                    list_1[1], list_1[2], list_1[3], list_1[4]), html.Br(), 'Пересохранить существующие данные'
                  
                return no_update, is_open, not is_open2, text

            columslist = functions2.sql_data_LIMIT500(data = 'dash_pipes', limit = 1).columns
#             if not 'gilza' in dt.columns:
#                 dt['gilza'] = None
            dt['gilza'] = None
                
            dt['id']=dt['diametrtruby'].apply(lambda x: format(float(x), '.2f'))+'_'+\
                    dt['stenkatruby'].apply(lambda x: format(float(x), '.2f'))+'_'+\
                    dt['diametropravkimmmpm'].apply(lambda x: format(float(x), '.2f'))+'_'+\
                    dt['kalibrelm'].apply(lambda x: format(float(x), '.2f'))+'_'+\
                    dt['markastali'].astype('str')
            dt=dt.replace('', 0)

            df_to_save = dt[columslist.tolist()]

            sss_heattreatment=''
            for i in range(0, df_to_save.shape[1]):
                sss_heattreatment=sss_heattreatment + "%s,"
            sss_heattreatment=sss_heattreatment[:-1]

            down_status = functions2.save_to_sql(sql_base = df_to_save, 
                                                         sss_heattreatment = sss_heattreatment, data = "dash_pipes",)
            # print(down_status)

            text ="Данные с диаметром: {}, стенкой: {}, калибром: {}, диаметром оправки: {} и маркой стали: {} сохранены".format(ff(dt['diametrtruby']),
                                    ff(dt['stenkatruby']), ff(dt['kalibrelm']),
                    ff(dt['diametropravkimmmpm']), ff(dt['markastali']),)
    
            if down_status == 'loaded':
                text ="Данные с диаметром: {}, стенкой: {}, калибром: {}, диаметром оправки: {} и маркой стали: {} сохранены".format(ff(dt['diametrtruby']),
                                    ff(dt['stenkatruby']), ff(dt['kalibrelm']), ff(dt['diametropravkimmmpm']), ff(dt['markastali']),)
            else:
                text ="Ошибка при сохранений данных. Попробуйте пересохранить еще раз"

            # print(text)
            return text, not is_open, is_open2, no_update
        
        else:
            text = 'Выберите трубу для сохранения'
            return text, not is_open, is_open2, no_update

    elif 'closeanalogmodal_page1' in changed_id: 
        return no_update, not is_open , is_open2, no_update
    elif 'rejectsave_page1' in changed_id: 
        return no_update, is_open , not is_open2, no_update
    
    elif 'oksave_page1' in changed_id: 
        text = 'check'
        # print(table_data[int(row_out[0])].get('Источник данных'))
        if table_data[int(row_out[0])].get('Источник данных') == 'Данные со стана':
            dt = functions2.data_by_ID(table_data[int(row_out[0])].get('Data'), 
                                  data = 'dash_analogue', id_ = "mpmin")
            dt = dt[dt['idpipesinuque'] == table_data[int(row_out[0])].get('ID')]
            
        elif table_data[int(row_out[0])].get('Источник данных') == 'Данные со стана (интеграция)':
            dt = functions2.data_by_ID(table_data[int(row_out[0])].get('ID'), data = 'integrazaya', id_ = "id")
                    
        elif table_data[int(row_out[0])].get('Источник данных') == 'Фактические настройки': #по источнику данных определяем нужну таблицу в Postgress
            dt = functions2.data_by_ID(table_data[int(row_out[0])].get('ID'), 
                              data = 'dash_pipes', id_ = "id")

        else:
            dt = functions2.data_by_ID(table_data[int(row_out[0])].get('ID'))
            dt.rename(columns={i:j for i,j in zip(old_colnames,new_colnames)}, inplace=True)

        columslist = functions2.sql_data_LIMIT500(data = 'dash_pipes', limit = 1).columns

        dt['gilza'] = None
        dt['id']=dt['diametrtruby'].apply(lambda x: format(float(x), '.2f'))+'_'+\
                    dt['stenkatruby'].apply(lambda x: format(float(x), '.2f'))+'_'+\
                    dt['diametropravkimmmpm'].apply(lambda x: format(float(x), '.2f'))+'_'+\
                    dt['kalibrelm'].apply(lambda x: format(float(x), '.2f'))+'_'+\
                    dt['markastali'].astype('str')
        dt=dt.replace('', 0)
        df_to_save = dt[columslist.tolist()]
        #delete duplicate###
        functions2.delete_row_by_ID(del_id=ff(dt['id']), data = "dash_pipes", id_ = 'id')

        sss_heattreatment=''
        for i in range(0, df_to_save.shape[1]):
            sss_heattreatment=sss_heattreatment + "%s,"
        sss_heattreatment=sss_heattreatment[:-1]

        down_status = functions2.save_to_sql(sql_base = df_to_save, 
                                                         sss_heattreatment = sss_heattreatment, data = "dash_pipes",)
        # print(down_status)
   
        if down_status == 'loaded':
            text ="Данные с диаметром: {}, стенкой: {}, калибром: {}, диаметром оправки: {} и маркой стали: {} пересохранены".format(ff(dt['diametrtruby']),
                                    ff(dt['stenkatruby']), ff(dt['kalibrelm']), ff(dt['diametropravkimmmpm']), ff(dt['markastali']),)
        else:
            text ="Ошибка при сохранений данных. Попробуйте пересохранить еще раз"

        # print(text)
        return text, not is_open, not is_open2, no_update


###CALBACK FROM FEDOROV FEDOR   
    
####################################################################################################
@app.callback(
    ###### Output
    # table_1,table_2,plot_1
    Output('postgres_datatable', 'children'),
    Output('postgres_datatable2', 'children'),
    Output('out-plot', 'children'),
    # style for table_1,table_2,plot_1
    Output('postgres_datatable', 'style'),
    Output('postgres_datatable2', 'style'),
    Output('out-plot', 'style'),
    # clicks to save_to_postgres
    Output('save_to_postgres', 'n_clicks'),
    # clicks to del_to_postgres
    Output('del_to_postgres', 'n_clicks'),
    # clicks to add rows
    Output('editing-rows-button', 'n_clicks'),
    # text
    Output('placeholder', 'children'),
    # index
    Output('index', 'n_clicks'),
    # editable tabl_1
    Output('out-table', 'editable'),
    # текст кнопки add
    Output('editing-rows-button', 'children'),
    Output('out-table', 'selected_rows'),
    # # отмена выбора
    # Output('cancel', 'n_clicks'),
    ###### Input
    # add rows
    [Input('editing-rows-button', 'n_clicks'),
    # save_to_postgres
    Input('save_to_postgres', 'n_clicks'),
    # save_to_postgres
    Input('del_to_postgres', 'n_clicks'),
    # number row
    # Input('out-table', 'derived_virtual_selected_rows'),
    Input('out-table', 'derived_viewport_selected_row_ids'),
    Input("interval", "n_intervals"),
    # index
    Input("index", "n_clicks")],
    ###### State
    # table_1
    State('out-table', 'data'),
    # table_1
    State('out-table2', 'data'),
    prevent_initial_call=True)
def df_to_csv_0(add_n_clicks, n_clicks,del_clicks,row, n_intervals, index,dataset,dataset2): 
    # print(add_n_clicks)
    # print(n_clicks)
    # print(del_clicks)
    # print(row)
    # print(n_intervals)
    # print(index)
    # print(cancel)
    # print(dataset)
    # print(dataset2) 
    ##############
    pgg = pd.DataFrame(dataset)
    #print(len(pgg))
    # первоначальный вывод таблицы #row==[]
    if add_n_clicks==0 and n_clicks==0 and del_clicks==0 and row==[] and n_intervals==None and dataset==[] and dataset2==[]:
        #print('условие 1')
        return table_1(functions.read_sql_ff('dash_pipes'),row=[],data_name=data_name)[0],\
        no_update,\
        no_update,\
        {'display': 'block'},\
        no_update,\
        no_update,\
        no_update,\
        no_update,\
        no_update,\
        no_update,\
        no_update,\
        True,\
        no_update,\
        no_update
    # сохранение после добавления строк
    #elif add_n_clicks>-999999 and n_clicks>0 and del_clicks>-999999 and row==[] and n_intervals==None and dataset!=[] and dataset2==[]:
    elif add_n_clicks>-999999 and n_clicks>0 and del_clicks>-999999 and row==[] and n_intervals==None and dataset!=[]:
    
        #print('условие 2')
        n_clicks=0
        pg = pd.DataFrame(dataset)
        pysto=0
        param=['diametropravkimmmpm','diametrtruby','stenkatruby','markastali','kalibrelm']
        for i in param:
            pysto+=int(pg[pg[i]==''].shape[0])
        if pysto==0:
            pg['id']=pg['diametrtruby'].apply(lambda x: format(float(x), '.2f'))+'_'+\
                pg['stenkatruby'].apply(lambda x: format(float(x), '.2f'))+'_'+\
                pg['diametropravkimmmpm'].apply(lambda x: format(float(x), '.2f'))+'_'+\
                pg['kalibrelm'].apply(lambda x: format(float(x), '.2f'))+'_'+\
                pg['markastali'].astype('str')
            pg=pg.replace('', 0)
            for i in pg.columns:
                if i=='markastali' or i=='id':
                    pg[i]=pg[i].astype('object')
                else:
                    pg[i]=pg[i].astype('float')
            ################# проверка на id
            data_count=pg[['id','diametropravkimmmpm','diametrtruby','stenkatruby','markastali','kalibrelm','gilza']].groupby(['id','diametropravkimmmpm','diametrtruby','stenkatruby','markastali','kalibrelm'],as_index=False).agg('count')
            data_count.sort_values(by='gilza',ascending=False,inplace=True)
            if data_count[data_count['gilza']>1].shape[0]!=0:
                #print('условие 2.1')
                data_count.reset_index(drop=True,inplace=True)
                to_text=data_count[data_count['gilza']>1].loc[0,'id']
                return table_1(functions.read_sql_ff('dash_pipes'),row=[],data_name=data_name)[0],\
                    no_update,\
                    no_update,\
                    {'display': 'block','editable':True},\
                    no_update,\
                    no_update,\
                    n_clicks,\
                    no_update,\
                    no_update,\
                    html.Plaintext('данные с диаметром: '+to_text.split('_')[0]+', стенкой: '+to_text.split('_')[1]+', диаметром оправки: '+to_text.split('_')[2]+', калибром: '+to_text.split('_')[3]+', маркой стали: '+to_text.split('_')[4]+' уже есть в данных',style={'color': 'green', 'font-weight': 'bold', 'font-size': 'large'}),\
                    no_update,\
                    True,\
                    no_update,\
                    no_update
            else:
                # print('условие 2.2')
                # print(len(pg))
                #functions.update_sql_ff(pg,'dash_pipes')
                functions.update_sql_ff(pg.drop_duplicates(keep='last'),'dash_pipes')
                
                return table_1(functions.read_sql_ff('dash_pipes'),row=[],data_name=data_name)[0],\
                no_update,\
                no_update,\
                {'display': 'block','editable':True},\
                no_update,\
                no_update,\
                n_clicks,\
                no_update,\
                no_update,\
                html.Plaintext("данные сохранены "+str(datetime.now() + timedelta(hours=3)),style={'color': 'green', 'font-weight': 'bold', 'font-size': 'large'}),\
                no_update,\
                True,\
                no_update,\
                no_update
        else:
            return no_update,\
                no_update,\
                no_update,\
                {'display': 'block','editable':True},\
                no_update,\
                no_update,\
                n_clicks,\
                no_update,\
                no_update,\
                html.Plaintext('заполните все поля',style={'color': 'green', 'font-weight': 'bold', 'font-size': 'large'}),\
                no_update,\
                True,\
                no_update,\
                no_update
    # выбoр по активной строке
    elif add_n_clicks>-999999 and n_clicks==0 and del_clicks>-999999 and row!=[] and n_intervals==None and dataset!=[] and dataset2!=-999999 and list(pgg[pgg['id']==row[0]].index)[0]!=index:
        #print('условие 3')
        # index=row[0]
        pg = pd.DataFrame(dataset)
        #print(pg)
        index=list(pg[pg['id']==row[0]].index)[0]
        pysto=0
        param=['diametropravkimmmpm','diametrtruby','stenkatruby','markastali','kalibrelm']
        for i in param:
            pysto+=int(pg[pg[i]==''].shape[0])
        if pysto==0:
            pg['id']=pg['diametrtruby'].apply(lambda x: format(float(x), '.2f'))+'_'+\
                pg['stenkatruby'].apply(lambda x: format(float(x), '.2f'))+'_'+\
                pg['diametropravkimmmpm'].apply(lambda x: format(float(x), '.2f'))+'_'+\
                pg['kalibrelm'].apply(lambda x: format(float(x), '.2f'))+'_'+\
                pg['markastali'].astype('str')
            pg=pg.replace('', 0)
            for i in pg.columns:
                if i=='markastali' or i=='id':
                    pg[i]=pg[i].astype('object')
                else:
                    pg[i]=pg[i].astype('float')
            ################# проверка на id
            data_count=pg[['id','diametropravkimmmpm','diametrtruby','stenkatruby','markastali','kalibrelm','gilza']].groupby(['id','diametropravkimmmpm','diametrtruby','stenkatruby','markastali','kalibrelm'],as_index=False).agg('count')
            data_count.sort_values(by='gilza',ascending=False,inplace=True)
            if data_count[data_count['gilza']>1].shape[0]!=0:
                data_count.reset_index(drop=True,inplace=True)
                to_text=data_count[data_count['gilza']>1].loc[0,'id']
                return no_update,\
                no_update,\
                no_update,\
                {'display': 'block'},\
                no_update,\
                no_update,\
                no_update,\
                no_update,\
                no_update,\
                html.Plaintext('данные с диаметром: '+to_text.split('_')[0]+', стенкой: '+to_text.split('_')[1]+', диаметром оправки: '+to_text.split('_')[2]+', калибром: '+to_text.split('_')[3]+', маркой стали: '+to_text.split('_')[4]+' уже есть в данных',style={'color': 'green', 'font-weight': 'bold', 'font-size': 'large'}),\
                index,\
                False,\
                'Отмена выбранной строки',\
                no_update
            else:
                table_2_plot_2=table_1(pd.DataFrame(dataset).replace('', 0),selected_rows=list(pg[pg['id']==row[0]].index),row=list(pg[pg['id']==row[0]].index),data_name=data_name)
                try:
                    fedorovfedor=pd.DataFrame(table_2_plot_2[1])
                    return no_update,\
                        table_2_plot_2[1],\
                        table_2_plot_2[2],\
                        {'display': 'block'},\
                        {'display': 'block'},\
                        {'display': 'block'},\
                        no_update,\
                        no_update,\
                        no_update,\
                        [],\
                        index,\
                        False,\
                        'Отмена выбранной строки',\
                        no_update
                except:
                    return table_1(pd.DataFrame(dataset),row=[],data_name=data_name)[0],\
                        no_update,\
                        no_update,\
                        {'display': 'block'},\
                        no_update,\
                        no_update,\
                        no_update,\
                        no_update,\
                        no_update,\
                        html.Plaintext('Сначала сохраните данные',style={'color': 'red', 'font-weight': 'bold', 'font-size': 'large'}),\
                        index,\
                        False,\
                        no_update,\
                        no_update
        else:
            return no_update,\
                no_update,\
                no_update,\
                {'display': 'block'},\
                no_update,\
                no_update,\
                no_update,\
                no_update,\
                no_update,\
                html.Plaintext('сначала заполните все пустые поля',style={'color': 'green', 'font-weight': 'bold', 'font-size': 'large'}),\
                index,\
                False,\
                'Отмена выбранной строки',\
                no_update
    # добавление строки
#     elif add_n_clicks>0 and add_n_clicks!=-999999 and n_clicks>-999999 and del_clicks>-999999\
#     and row==[] and n_intervals==None and dataset!=[] and dataset2==[]:
    elif add_n_clicks>0 and add_n_clicks!=-999999 and n_clicks>-999999 and del_clicks>-999999 and row==[] and n_intervals==None and dataset!=[]:
        #print('условие 4')
        add_n_clicks=0
        pg = pd.DataFrame(dataset)
        pysto=0
        param=['diametropravkimmmpm','diametrtruby','stenkatruby','markastali','kalibrelm']
        for i in param:
            pysto+=int(pg[pg[i]==''].shape[0])
        if pysto==0:
            pg['id']=pg['diametrtruby'].apply(lambda x: format(float(x), '.2f'))+'_'+\
                pg['stenkatruby'].apply(lambda x: format(float(x), '.2f'))+'_'+\
                pg['diametropravkimmmpm'].apply(lambda x: format(float(x), '.2f'))+'_'+\
                pg['kalibrelm'].apply(lambda x: format(float(x), '.2f'))+'_'+\
                pg['markastali'].astype('str')
            pg=pg.replace('', 0)
            for i in pg.columns:
                if i=='markastali' or i=='id':
                    pg[i]=pg[i].astype('object')
                else:
                    pg[i]=pg[i].astype('float')
            data_count=pg[['id','diametropravkimmmpm','diametrtruby','stenkatruby','markastali','kalibrelm','gilza']].groupby(['id','diametropravkimmmpm','diametrtruby','stenkatruby','markastali','kalibrelm'],as_index=False).agg('count')
            data_count.sort_values(by='gilza',ascending=False,inplace=True)
            if data_count[data_count['gilza']>1].shape[0]!=0:
                data_count.reset_index(drop=True,inplace=True)
                to_text=data_count[data_count['gilza']>1].loc[0,'id']
                return no_update,\
                no_update,\
                no_update,\
                {'display': 'block'},\
                no_update,\
                no_update,\
                no_update,\
                no_update,\
                add_n_clicks,\
                html.Plaintext('данные с диаметром: '+to_text.split('_')[0]+', стенкой: '+to_text.split('_')[1]+', диаметром оправки: '+to_text.split('_')[2]+', калибром: '+to_text.split('_')[3]+', маркой стали: '+to_text.split('_')[4]+' уже есть в данных',style={'color': 'green', 'font-weight': 'bold', 'font-size': 'large'}),\
                no_update,\
                True,\
                no_update,\
                no_update
            else:
                dataset.append(dict(zip(list(data_name['name_bd']),[''] * len(list(data_name['name_bd'])))))
                dataset=[dataset[-1]]+dataset[:-1]
                return table_1(pd.DataFrame(dataset),row=[],data_name=data_name)[0],\
                    no_update,\
                    no_update,\
                    {'display': 'block'},\
                    no_update,\
                    no_update,\
                    no_update,\
                    no_update,\
                    add_n_clicks,\
                    no_update,\
                    no_update,\
                    True,\
                    no_update,\
                    no_update
        else:
            return no_update,\
            no_update,\
            no_update,\
            {'display': 'block'},\
            no_update,\
            no_update,\
            no_update,\
            no_update,\
            add_n_clicks,\
            html.Plaintext('сначала заполните все пустые поля',style={'color': 'green', 'font-weight': 'bold', 'font-size': 'large'}),\
            no_update,\
            True,\
            no_update,\
            no_update
    # отмена выбора
    elif add_n_clicks>0 and n_clicks>-999999 and del_clicks>-999999 and row!=[] and n_intervals==None and dataset!=[] and dataset2!=-999999:
        #print('условие  5')
        add_n_clicks=0
        index=-1
        return no_update,\
            no_update,\
            no_update,\
            {'display': 'block'},\
            {'display': 'none'},\
            {'display': 'none'},\
            no_update,\
            no_update,\
            add_n_clicks,\
            [],\
            index,\
            True,\
            'Добавить',\
            []
    # сохранение 2-х табличек
    elif add_n_clicks>-999999 and n_clicks>0 and del_clicks>-999999 and row!=[] and n_intervals==None and dataset!=[] and dataset2!=[]:
    #elif add_n_clicks>-999999 and n_clicks>0 and del_clicks>-999999 and n_intervals==None and dataset!=[] and dataset2!=[]:

        if row[0] == '':
            #print('ok')
            dataset2_1 = [
             {'стан': 'МПМ', 'Номер клети': 1, 'Диаметр Валков': 0, 'Обороты': 0, 'Зазоры': 0}, 
             {'стан': 'МПМ', 'Номер клети': 2, 'Диаметр Валков': 0, 'Обороты': 0, 'Зазоры': 0}, 
             {'стан': 'МПМ', 'Номер клети': 3, 'Диаметр Валков': 0, 'Обороты': 0, 'Зазоры': 0}, 
             {'стан': 'МПМ', 'Номер клети': 4, 'Диаметр Валков': 0, 'Обороты': 0, 'Зазоры': 0}, 
             {'стан': 'МПМ', 'Номер клети': 5, 'Диаметр Валков': 0, 'Обороты': 0, 'Зазоры': 0}, 
             {'стан': 'МПМ', 'Номер клети': 6, 'Диаметр Валков': 0, 'Обороты': 0, 'Зазоры': 0}, 
             {'стан': 'МПМ', 'Номер клети': 7, 'Диаметр Валков': 0, 'Обороты': 0, 'Зазоры': 0}, 
             {'стан': 'ИКС', 'Номер клети': 1, 'Диаметр Валков': 0, 'Обороты': 0, 'Зазоры': 0}, 
             {'стан': 'ИКС', 'Номер клети': 2, 'Диаметр Валков': 0, 'Обороты': 0, 'Зазоры': 0}, 
             {'стан': 'ИКС', 'Номер клети': 3, 'Диаметр Валков': 0, 'Обороты': 0, 'Зазоры': 0}, 
             {'стан': 'ИКС', 'Номер клети': 4, 'Диаметр Валков': 0, 'Обороты': 0, 'Зазоры': 0}, 
             {'стан': 'ИКС', 'Номер клети': 5, 'Диаметр Валков': 0, 'Обороты': 0, 'Зазоры': 0}, 
             {'стан': 'ИКС', 'Номер клети': 6, 'Диаметр Валков': 0, 'Обороты': 0, 'Зазоры': 0}, 
             {'стан': 'ИКС', 'Номер клети': 7, 'Диаметр Валков': 0, 'Обороты': 0, 'Зазоры': 0}, 
             {'стан': 'ИКС', 'Номер клети': 8, 'Диаметр Валков': 0, 'Обороты': 0, 'Зазоры': 0}, 
             {'стан': 'ИКС', 'Номер клети': 9, 'Диаметр Валков': 0, 'Обороты': 0, 'Зазоры': 0}]
        else:
            dataset2_1 = dataset2
        #print('условие 6')
        n_clicks=0
        text = df_to_csv_2(list(pgg[pgg['id']==row[0]].index),dataset,dataset2_1)
        # print(text)
        table_2_plot_2=table_1(functions.read_sql_ff('dash_pipes'),selected_rows=list(pgg[pgg['id']==row[0]].index),row=list(pgg[pgg['id']==row[0]].index),data_name=data_name)
        
        return table_2_plot_2[0],\
        table_2_plot_2[1],\
        table_2_plot_2[2],\
        no_update,\
        no_update,\
        no_update,\
        n_clicks,\
        no_update,\
        no_update,\
        text,\
        no_update,\
        False,\
        no_update,\
        no_update
    # таблица 1
    # таблица 2
    # график 1
    # видимость табл 1
    # видимость табл 2
    # видимость графика 1
    # кол-во  нажатий на сохранить
    # кол-во  нажатий на del
    # кол-во нажатий на add rows
    # текст
    # невидимая кнопка с выбором строки
    # редактирование табл 1
    # текст кнапки  add
    # отмена выбранной строки

    # удаление
    elif add_n_clicks>-999999 and n_clicks>-999999 and del_clicks>0 and row!=[] and n_intervals==None and dataset!=[] and dataset2!=[]:
        del_clicks=0
        pg = pd.DataFrame(dataset)
        # print(ff(pg['id'][list(pg[pg['id']==row[0]].index)]))
        functions2.delete_row_by_ID(del_id=ff(pg['id'][list(pg[pg['id']==row[0]].index)]), data = "dash_pipes", id_ = 'id')
        # print('условие 7  - удаление')
        index = -1
        return table_1(functions.read_sql_ff('dash_pipes'),row=[],data_name=data_name)[0],\
        no_update,\
        no_update,\
        no_update,\
        {'display': 'none'},\
        {'display': 'none'},\
        no_update,\
        del_clicks,\
        no_update,\
        [],\
        index,\
        no_update,\
        'Добавить',\
        []
    elif add_n_clicks>-999999 and n_clicks>-999999 and del_clicks>0 and row!=[] and n_intervals==None and dataset!=[] and dataset2==[]:
        del_clicks=0
        pg = pd.DataFrame(dataset)
        functions2.delete_row_by_ID(del_id=ff(pg['id'][list(pg[pg['id']==row[0]].index)]), data = "dash_pipes", id_ = 'id')
        # print('условие 7.2  - удаление')
        index = -1
        return table_1(functions.read_sql_ff('dash_pipes'),row=[],data_name=data_name)[0],\
        no_update,\
        no_update,\
        no_update,\
        {'display': 'none'},\
        {'display': 'none'},\
        no_update,\
        del_clicks,\
        no_update,\
        [],\
        index,\
        no_update,\
        'Добавить',\
        []
    else:
        # print('else')
        return no_update,\
        no_update,\
        no_update,\
        no_update,\
        no_update,\
        no_update,\
        no_update,\
        no_update,\
        no_update,\
        no_update,\
        no_update,\
        no_update,\
        no_update,\
        no_update
####################################################################################################    

###calback resolve problems with filter on 2 and other pages###
@app.callback(
    Output('out-table', "page_current"),
    Input('out-table', "filter_query"))
def update_table_filter_3(filter):
    return 0  

@app.callback(
    Output('real_pipes', "page_current"),
    Input('real_pipes', "filter_query"))
def update_table_filter_4(filter):
    #print(filter)
    return 0 


@app.callback(Output('updatetime', 'children'), #display last update time
             Output('page3_maintable', 'children'), #display default table
             Output('real_pipes', 'selected_rows'), #unselect selected row
             Output('modal_page3', 'is_open'), #modal for chosing pipes by condition
             Output('selector_kalibrelm', 'options'), #unique options for dropdown
             Output('selector_markastali', 'options'), #unique options for dropdown
             Output('selector_diametrtruby', 'options'), #unique options for dropdown
             Output('selector_stenkatruby', 'options'), #unique options for dropdown
             Output('selector_diametropravkimmmpm', 'options'), #unique options for dropdown
            #graphs which emerged when selected row
             Output('page3_table', 'children'), #table1
             Output('page3_graph', 'children'), #graph speed
             Output('toki_mpm', 'children'), #graph current
             Output('toki_esm', 'children'), #graph current
             Input('renew_data', 'n_clicks'),
             Input('condition_load_btn', 'n_clicks'),
             Input('ok_page3', 'n_clicks'),
             Input('return_page3', 'n_clicks'),
             Input('selector_kalibrelm', 'value'),
             Input('selector_markastali', 'value'),
             Input('selector_diametrtruby', 'value'),
             Input('selector_stenkatruby', 'value'),
             Input('selector_diametropravkimmmpm', 'value'),
             Input('date_from', 'value'),
             Input('date_to', 'value'),
             #Input('real_pipes', 'selected_rows'),
             Input('real_pipes', 'derived_viewport_selected_row_ids'),
             State('modal_page3', 'is_open'),
             prevent_initial_call=True)
def page3update(n, selectdata, ok, refuse,  kalibrelm_val, markastali_val, diametrtruby_val, stenkatruby_val, diametropravkimmmpm_val, date_from, date_to, row_id, is_open):
    changed_id = [p['prop_id'] for p in dash.callback_context.triggered][0]
    query_date_to = (datetime.now() + timedelta(hours=3)).strftime("%Y-%m-%d %H:%M:%S")
    query_date_from = '1990-01-00 00:00:00'
    if 'renew_data' in changed_id:

        return html.P('Обновлено:  ' + str((datetime.now() + timedelta(hours=3)).strftime("%d-%m-%Y %H:%M:%S"))), table_realpipes(functions2.sql_data_LIMIT500(data = 'integrazaya', order_id = 'timestamp', limit = 10), 0)[0], [], no_update, no_update, no_update, no_update, no_update, no_update, None, None, None, None
    
    elif 'condition_load_btn' in changed_id:
        
        #upload unique variables for dropdowns from Postgress
        check = functions2.sql_data_select(vals = 'kalibrelm', data = 'integrazaya')
        kalibrelm = [float(ff(n)) for n in check]    
        
        check = functions2.sql_data_select(vals = 'diametropravkimmmpm', data = 'integrazaya')
        diametropravkimmmpm = [float(ff(n)) for n in check]

        check = functions2.sql_data_select(vals = 'diametrtruby', data = 'integrazaya')
        diametrtruby = [float(ff(n)) for n in check]
        
        check = functions2.sql_data_select(vals = 'stenkatruby', data = 'integrazaya')
        stenkatruby = [float(ff(n)) for n in check]
        
        check = functions2.sql_data_select(vals = 'markastali', data = 'integrazaya')
        markastali = [ff(n) for n in check]
        
        return no_update, no_update, no_update, not is_open, [{'label': i, 'value': i} for i in kalibrelm], [{'label': i, 'value': i} for i in markastali], [{'label': i, 'value': i} for i in diametrtruby], [{'label': i, 'value': i} for i in stenkatruby], [{'label': i, 'value': i} for i in diametropravkimmmpm], no_update, no_update, no_update, no_update
    elif 'return_page3' in changed_id: #undo query for pipes
        return no_update, no_update, no_update, not is_open, no_update, no_update, no_update, no_update, no_update, no_update, no_update, no_update, no_update
    elif 'ok_page3' in changed_id: #pipes query by condition
        
        #create query
        query_PSQ = query_kalibrelm = query_diametrtruby = query_stenkatruby = query_diametropravkimmmpm = query_markastali =''
        sql_list = []

        if kalibrelm_val and len(kalibrelm_val) > 0:
            query_kalibrelm = 'kalibrelm = ANY (%s) AND '
            sql_list.append(kalibrelm_val)
        if diametrtruby_val and len(diametrtruby_val) > 0:
            query_diametrtruby = 'diametrtruby = ANY (%s) AND '
            sql_list.append(diametrtruby_val)
        if stenkatruby_val and len(stenkatruby_val) > 0:
            query_stenkatruby = 'stenkatruby = ANY (%s) AND ' 
            sql_list.append(stenkatruby_val)
        if diametropravkimmmpm_val and len(diametropravkimmmpm_val) > 0:
            query_diametropravkimmmpm = 'diametropravkimmmpm = ANY (%s) AND '
            sql_list.append(diametropravkimmmpm_val)
        if markastali_val and len(markastali_val) > 0:
            query_markastali = 'markastali = ANY (%s) AND '
            sql_list.append(markastali_val)
            
        if date_from and date_from != []:
            query_date_from = datetime.fromisoformat(date_from).strftime("%Y-%m-%d %H:%M:%S")
        sql_list.append(query_date_from)
        
        if date_to and date_to != []:
            #delta = timedelta.Timedelta(hours=23, minutes = 59, seconds = 59)
            delta = timedelta(hours=23, minutes = 59, seconds = 59)
            query_date_to = (datetime.fromisoformat(date_to) + delta).strftime("%Y-%m-%d %H:%M:%S")
        sql_list.append(query_date_to)

        query_PSQ = 'WHERE ' + query_kalibrelm + query_diametrtruby + query_stenkatruby + query_diametropravkimmmpm + query_markastali
        query_PSQ = query_PSQ + 'timestamp BETWEEN (%s) and (%s) ORDER BY timestamp DESC LIMIT 500'

        return html.P('Обновлено:  ' + str((datetime.now() + timedelta(hours=3)).strftime("%d-%m-%Y %H:%M:%S"))), table_realpipes(functions2.data_query(query_PSQ, sql_list), 0)[0], [], not is_open, no_update, no_update, no_update, no_update, no_update, None, None, None, None
    
    elif 'real_pipes' in changed_id:
        if row_id and row_id != []:
            # print(row_id)
            currents = functions2.data_by_ID(track_id = ff(row_id), data = 'current', id_ = "id") #currents
            df_by_id = functions2.data_by_ID(track_id = ff(row_id), data = 'integrazaya', id_ = "id") #rows data

            graph_1 = graph_2 = None
            
            if currents_graph(currents)[0]:
                graph_1 = dcc.Graph(figure = currents_graph(currents)[0], style = {'height': '400px'})
            if currents_graph(currents)[1]:
                graph_2 = dcc.Graph(figure = currents_graph(currents)[1], style = {'height': '400px'})
            return no_update, no_update, no_update, is_open, no_update, no_update, no_update, no_update, no_update, table_realpipes(df_by_id, 0)[1], table_realpipes(df_by_id, 0)[2], graph_1, graph_2
       
        return no_update, no_update, no_update, is_open, no_update, no_update, no_update, no_update, no_update, no_update, no_update, no_update, no_update
    else:
        return no_update, no_update, no_update, is_open, no_update, no_update, no_update, no_update, no_update, no_update, no_update, no_update, no_update

###save as ideal pipes###
@app.callback(Output('load_result2', 'children'), #display last update time
             Output('modal_analog_page3', 'is_open'),
             Output('modal_analog2_page3', 'is_open'),
             Output('load_result', 'children'),
             Input('save_ideal_btn', 'n_clicks'),
             Input('closeanalogmodal_page3', 'n_clicks'),
             Input('rejectsave_page3', 'n_clicks'),
             Input('oksave_page3', 'n_clicks'),
             State('real_pipes', 'derived_viewport_selected_row_ids'),
             State('modal_analog_page3', 'is_open'),
             State('modal_analog2_page3', 'is_open'),
             prevent_initial_call=True)
def page3saveideal(btn_save, close, reject, save, row_id, is_open , is_open2):
    # print('save') 
    # print(row_id)
    changed_id = [p['prop_id'] for p in dash.callback_context.triggered][0]
    if 'save_ideal_btn' in changed_id:
        if row_id and row_id != []:
            df_by_id = functions2.data_by_ID(track_id = ff(row_id), 
                                             data = 'integrazaya', id_ = "id")[['diametrtruby', 'stenkatruby', 'kalibrelm', 
                                                   'diametropravkimmmpm' ,'markastali']]

            list_check = ff(df_by_id.values.tolist()) 
            list_1 = [float(i) for i in list_check[0:4]]
            list_1.append(list_check[4]) 
            
            ####проверка на наличие####
            sql_query2 = 'WHERE diametrtruby = (%s) AND stenkatruby = (%s) AND  kalibrelm = (%s) AND diametropravkimmmpm = (%s) AND markastali = (%s) ORDER BY ctid DESC LIMIT 500'
            dfquery = functions2.data_query(sql_query2, list_1,  data = 'dash_pipes')
            
            if len(dfquery) > 0:
                text ="Данные с диаметром: {}, стенкой: {}, калибром: {}, диаметром оправки: {} и маркой стали: {} уже существуют".format(list_1[0],
                                                                    list_1[1], list_1[2], list_1[3], list_1[4]), html.Br(), 'Пересохранить существующие данные'
                  
                
                return no_update, is_open , not is_open2, text
            else:

                df_by_id = functions2.data_by_ID(track_id = ff(row_id), 
                                             data = 'integrazaya', id_ = "id")
                df_by_id['id']=df_by_id['diametrtruby'].apply(lambda x: format(float(x), '.2f'))+'_'+\
                df_by_id['stenkatruby'].apply(lambda x: format(float(x), '.2f'))+'_'+\
                df_by_id['diametropravkimmmpm'].apply(lambda x: format(float(x), '.2f'))+'_'+\
                df_by_id['kalibrelm'].apply(lambda x: format(float(x), '.2f'))+'_'+\
                df_by_id['markastali'].astype('str')
                df_by_id=df_by_id.replace('', 0)

                #functions2.delete_row_by_ID(del_id=ff(df_by_id['id']), data = "dash_pipes", id_ = 'id')

                columslist = functions2.sql_data_LIMIT500(data = 'dash_pipes', limit = 1).columns
                df_by_id['gilza'] = None
                
                df_to_save = df_by_id[columslist.tolist()]

                sss_heattreatment=''
                for i in range(0, df_to_save.shape[1]):
                    sss_heattreatment=sss_heattreatment + "%s,"
                sss_heattreatment=sss_heattreatment[:-1]

                down_status = functions2.save_to_sql(sql_base = df_to_save, 
                                                                 sss_heattreatment = sss_heattreatment, data = "dash_pipes",)
                # print(down_status)
                if down_status == 'loaded':
                    text ="Данные с диаметром: {}, стенкой: {}, калибром: {}, диаметром оправки: {} и маркой стали: {} сохранены".format(ff(df_by_id['diametrtruby']), 
                    ff(df_by_id['stenkatruby']), ff(df_by_id['kalibrelm']), ff(df_by_id['diametropravkimmmpm']), ff(df_by_id['markastali']),)
                else:
                    text ="Ошибка при сохранений данных. Попробуйте пересохранить еще раз1"
                    # print(text)
                
                return text,  not is_open , is_open2, no_update
            
#             return no_update, is_open , not is_open2
        else:
            return 'Выберите трубу для сохранения', not is_open , is_open2, no_update
        
    elif 'oksave_page3' in changed_id: 
        df_by_id = functions2.data_by_ID(track_id = ff(row_id), 
                                             data = 'integrazaya', id_ = "id") 
       
        
        df_by_id['id']=df_by_id['diametrtruby'].apply(lambda x: format(float(x), '.2f'))+'_'+\
                df_by_id['stenkatruby'].apply(lambda x: format(float(x), '.2f'))+'_'+\
                df_by_id['diametropravkimmmpm'].apply(lambda x: format(float(x), '.2f'))+'_'+\
                df_by_id['kalibrelm'].apply(lambda x: format(float(x), '.2f'))+'_'+\
                df_by_id['markastali'].astype('str')
        df_by_id=df_by_id.replace('', 0)
        
        functions2.delete_row_by_ID(del_id=ff(df_by_id['id']), data = "dash_pipes", id_ = 'id')
        
        columslist = functions2.sql_data_LIMIT500(data = 'dash_pipes', limit = 1).columns
        df_by_id['gilza'] = None
        df_to_save = df_by_id[columslist.tolist()]
        
        sss_heattreatment=''
        for i in range(0, df_to_save.shape[1]):
            sss_heattreatment=sss_heattreatment + "%s,"
        sss_heattreatment=sss_heattreatment[:-1]

        down_status = functions2.save_to_sql(sql_base = df_to_save, 
                                                     sss_heattreatment = sss_heattreatment, data = "dash_pipes",)
        # print(down_status)

        text ="Данные с диаметром: {}, стенкой: {}, калибром: {}, диаметром оправки: {} и маркой стали: {} пересохранены".format(ff(df_by_id['diametrtruby']),
                                ff(df_by_id['stenkatruby']), ff(df_by_id['kalibrelm']),
                ff(df_by_id['diametropravkimmmpm']), ff(df_by_id['markastali']),)
    
        if down_status == 'loaded':
            text ="Данные с диаметром: {}, стенкой: {}, калибром: {}, диаметром оправки: {} и маркой стали: {} сохранены".format(ff(df_by_id['diametrtruby']),
                                ff(df_by_id['stenkatruby']), ff(df_by_id['kalibrelm']), ff(df_by_id['diametropravkimmmpm']), ff(df_by_id['markastali']),)
        else:
            text ="Ошибка при сохранений данных. Попробуйте пересохранить еще раз"

        return text,  not is_open , not is_open2, no_update
        
    elif 'closeanalogmodal_page3' in changed_id: 
        return no_update, not is_open , is_open2, no_update
    elif 'rejectsave_page3' in changed_id: 
        return no_update, is_open , not is_open2, no_update
        
# Update the index
@app.callback(Output('page-content', 'children'),
              [Input('url', 'pathname')])
def display_page(pathname):
    if pathname == '/':
        return page_1_layout
#     elif pathname == '/model':
#         return page_1_layout
    elif pathname == '/table':
        return page_2_layout
    elif pathname == '/realpipes':
        page_3_layout = dbc.Container([
            dcc.Interval(id='interval_pg3',  interval=60000, n_intervals=0),# interval update - 15 minites # interval=60000*15
            html.P(),
            # dbc.Row([dbc.Col(group3, width = 10), 
            # dbc.Col(html.P('Обновлено:  ' + str((datetime.now() + timedelta(hours=3)).strftime("%d-%m-%Y %H:%M:%S")), style = {'font-size': '10'}), 
            #                                             width = 2, id = 'updatetime')]),
            dbc.Row([dbc.Col(group3, width = 12)]),
            dbc.Row([dbc.Col(html.P('Обновлено:  ' + str((datetime.now() + timedelta(hours=3)).strftime("%d-%m-%Y %H:%M:%S")), style = {'font-size': '10'}), 
                                                        width = 12, id = 'updatetime')]),
            # html.P(),
            dbc.Row(
                    [
                      dbc.Col(
                    html.Div(table_realpipes(functions2.sql_data_LIMIT500(data = 'integrazaya', order_id = 'timestamp', limit = 10), 0)[0], id = 'page3_maintable'), 
                             width = 12)]),
            html.P(),
            dbc.Row(
                    [
                        dbc.Col(html.Div(id = 'page3_table'), width = 5),
                        dbc.Col(html.Div(id = 'page3_graph'), width = 7),

                    ]),
          dbc.Row([  
                        dbc.Col(html.Div(id = 'toki_mpm'), width = 6),
                        dbc.Col(html.Div(id = 'toki_esm'), width = 6)

                    ]),
        ],fluid = True)
        
        return page_3_layout
    else:
        return page_1_layout

if __name__ == '__main__':
    app.run_server(debug=False, host = '0.0.0.0', port = 8050)
